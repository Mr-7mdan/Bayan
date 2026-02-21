from __future__ import annotations

from fastapi import APIRouter, Depends, HTTPException, Query, Response, UploadFile, File, Form
from typing import Any
from uuid import uuid4
import os
import json
from sqlalchemy import inspect, text, or_
from urllib.parse import unquote
from sqlalchemy.orm import Session
from pydantic import BaseModel
try:
    import duckdb as _duckdb
except Exception:  # pragma: no cover
    _duckdb = None

from ..models import SessionLocal, Datasource, init_db, create_datasource, NewDatasourceInput, User, SyncTask, SyncState, SyncRun, SyncLock, DatasourceShare
from ..db import get_duckdb_engine, get_engine_from_dsn, run_sequence_sync, run_snapshot_sync, open_duck_native, get_active_duck_path
from ..api_ingest import run_api_sync
from ..db import dispose_engine_by_key, dispose_all_engines, dispose_duck_engine
from ..schemas import (
    DatasourceCreate,
    DatasourceOut,
    DatasourceDetailOut,
    DatasourceUpdate,
    DatasourceShareAddRequest,
    DatasourceShareOut,
    IntrospectResponse,
    SchemaInfo,
    TableInfo,
    ColumnInfo,
    SyncTaskCreate,
    SyncTaskOut,
    LocalStatsResponse,
    LocalTableStat,
    SyncRunOut,
    DatasourceExportItem,
    DatasourceImportItem,
    DatasourceImportRequest,
    DatasourceImportResponse,
    DatasourceImportResponse,
    DatasourceTransforms,
    TransformsPreviewRequest,
    PreviewResponse,
)
from ..sqlgen import build_sql
from ..config import settings
import re
from datetime import datetime, timezone
from ..security import encrypt_text, decrypt_text
from ..metrics import counter_inc
import logging
import os

router = APIRouter(prefix="/datasources", tags=["datasources"])

# local logger for sync runs (prints to console)
_log = logging.getLogger("app.sync")
if not _log.handlers:
    _h = logging.StreamHandler()
    _h.setFormatter(logging.Formatter("%(asctime)s %(levelname)s [%(name)s] %(message)s"))
    _log.addHandler(_h)
try:
    _log.setLevel(logging.INFO)
except Exception:
    pass
_log.propagate = False

DEBUG = str(os.getenv("API_SYNC_DEBUG", "1")).strip().lower() in ("1", "true", "on")


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def _http_for_db_error(e: Exception) -> HTTPException | None:
    try:
        msg = str(e) if e is not None else ""
        up = msg.upper()
        if ("HYT00" in up) or ("LOGIN TIMEOUT" in up):
            return HTTPException(status_code=504, detail="Database connectivity timeout (HYT00)")
        if ("08S01" in up) or ("TCP PROVIDER" in up):
            return HTTPException(status_code=502, detail="Database connection lost (08S01/TCP Provider)")
    except Exception:
        return None
    return None

@router.post("", response_model=DatasourceOut)
def create_ds(payload: DatasourceCreate, db: Session = Depends(get_db)):
    enc = encrypt_text(payload.connectionUri) if payload.connectionUri else None
    ds = create_datasource(
        db,
        NewDatasourceInput(
            name=payload.name,
            type=payload.type,
            connection_encrypted=enc,
            options=payload.options,
            user_id=payload.userId,
        ),
    )
    return DatasourceOut.model_validate(ds)


def _is_blackout_now(ds: Datasource) -> bool:
    """Returns True when current local server time falls within any blackout range.
    options_json: { "sync": { "blackoutDaily": [ {"start": "HH:MM", "end": "HH:MM"}, ... ] } }
    Ranges are inclusive of start, exclusive of end; supports wrap-around (e.g., 22:00 -> 06:00).
    """
    try:
        opts = json.loads(ds.options_json or "{}")
    except Exception:
        opts = {}
    # Optional hardening: validate Unpivot transforms for table scope
    try:
        from sqlalchemy import text as _text
        from ..db import get_engine_from_dsn, get_duckdb_engine
        from ..security import decrypt_text
        def _q_source(name: str) -> str:
            s = str(name or '').strip()
            if not s:
                return s
            d = (ds.type or '').lower()
            if 'mssql' in d or 'sqlserver' in d:
                parts = s.split('.')
                return '.'.join([p if (p.startswith('[') and p.endswith(']')) else f"[{p}]" for p in parts])
            if 'mysql' in d:
                parts = s.split('.')
                return '.'.join([p if (p.startswith('`') and p.endswith('`')) else f"`{p}`" for p in parts])
            return s
        def _list_cols_for_table(table_name: str) -> set[str]:
            try:
                if not ds.connection_encrypted:
                    engine = get_duckdb_engine()
                else:
                    dsn = decrypt_text(ds.connection_encrypted)
                    if not dsn:
                        return set()
                    engine = get_engine_from_dsn(dsn)
                with engine.connect() as conn:
                    if (ds.type or '').lower() in ("mssql", "mssql+pymssql", "mssql+pyodbc"):
                        probe = _text(f"SELECT TOP 0 * FROM {_q_source(table_name)} AS s")
                    else:
                        probe = _text(f"SELECT * FROM {_q_source(table_name)} WHERE 1=0")
                    res = conn.execute(probe)
                    return set([str(c) for c in res.keys()])
            except Exception:
                return set()
        def _norm(s: str) -> str:
            return (s or '').strip().strip('[]').strip('"').strip('`').split('.')[-1].lower()
        eff = payload.model_dump(by_alias=True)
        errors: list[str] = []
        for tr in (eff.get('transforms') or []):
            try:
                if str((tr or {}).get('type') or '').lower() != 'unpivot':
                    continue
                sc = (tr or {}).get('scope') or {}
                lvl = str(sc.get('level') or '').lower()
                if lvl == 'table' and sc.get('table'):
                    cols = _list_cols_for_table(str(sc.get('table')))
                    low = {_norm(c) for c in cols}
                    missing = [c for c in (tr.get('sourceColumns') or []) if _norm(c) not in low]
                    if missing:
                        errors.append(f"unpivot: unknown columns for table '{sc.get('table')}': {', '.join(missing)}")
                    kc = str(tr.get('keyColumn') or '').strip()
                    vc = str(tr.get('valueColumn') or '').strip()
                    if kc and _norm(kc) in low:
                        errors.append(f"unpivot: keyColumn collides with existing column '{kc}' on table '{sc.get('table')}'")
                    if vc and _norm(vc) in low:
                        errors.append(f"unpivot: valueColumn collides with existing column '{vc}' on table '{sc.get('table')}'")
            except Exception:
                continue
        if errors:
            raise HTTPException(status_code=400, detail='; '.join(errors))
    except HTTPException:
        raise
    except Exception:
        # Best-effort validation; ignore errors
        pass
    # (removed accidental validation block)
    blk = (((opts or {}).get("sync") or {}).get("blackoutDaily"))
    if not isinstance(blk, list):
        return False
    from datetime import datetime as _dt
    now = _dt.now()
    cur = now.hour * 60 + now.minute
    for w in blk:
        try:
            s = str((w or {}).get("start") or "00:00")
            e = str((w or {}).get("end") or "00:00")
            sh, sm = [int(x) for x in s.split(":")]
            eh, em = [int(x) for x in e.split(":")]
            start_m = sh * 60 + sm
            end_m = eh * 60 + em
            if start_m == end_m:
                # full-day blackout
                return True
            if start_m < end_m:
                # normal range
                if cur >= start_m and cur < end_m:
                    return True
            else:
                # wrap-around
                if cur >= start_m or cur < end_m:
                    return True
        except Exception:
            continue
    return False


def _max_concurrent(ds: Datasource) -> int:
    try:
        opts = json.loads(ds.options_json or "{}")
    except Exception:
        opts = {}
    v = (((opts or {}).get("sync") or {}).get("maxConcurrentQueries"))
    try:
        n = int(v)
        return max(1, n)
    except Exception:
        return 1


def _is_admin(db: Session, actor_id: str | None) -> bool:
    if not actor_id:
        return False
    u = db.query(User).filter(User.id == str(actor_id).strip()).first()
    return bool(u and (u.role or "user").lower() == "admin")


def _group_key_for(ds_id: str, schema: str | None, table: str, dest: str) -> str:
    sch = (schema or "public").strip()
    return f"{ds_id}:{sch}.{table}->{dest}"


def _task_to_out(db: Session, t: SyncTask) -> SyncTaskOut:
    st = db.query(SyncState).filter(SyncState.task_id == t.id).first()
    return SyncTaskOut(
        id=t.id,
        datasourceId=t.datasource_id,
        sourceSchema=t.source_schema,
        sourceTable=t.source_table,
        destTableName=t.dest_table_name,
        mode=t.mode,
        pkColumns=t.pk_columns,
        selectColumns=t.select_columns,
        sequenceColumn=t.sequence_column,
        batchSize=t.batch_size,
        scheduleCron=t.schedule_cron,
        enabled=t.enabled,
        customQuery=t.custom_query,
        groupKey=t.group_key,
        createdAt=t.created_at,
        lastRunAt=(st.last_run_at if st else None),
        lastRowCount=(st.last_row_count if st else None),
        inProgress=(st.in_progress if st else False),
        error=(st.error if st else None),
        progressCurrent=(st.progress_current if st else None),
        progressTotal=(st.progress_total if st else None),
        progressPhase=(st.progress_phase if st else None),
        startedAt=(st.started_at if st else None),
    )


@router.get("", response_model=list[DatasourceOut])
def list_ds(userId: str | None = Query(default=None), actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    if _is_admin(db, actorId) and (userId is None or (str(userId).strip().lower() in {"", "undefined", "null"})):
        q = db.query(Datasource)
        rows = q.order_by(Datasource.created_at.desc()).all()
        return [DatasourceOut.model_validate(r) for r in rows]
    if userId is None or (str(userId).strip().lower() in {"", "undefined", "null"}):
        uid = (actorId or "").strip()
    else:
        uid = str(userId).strip()
    if not uid:
        from sqlalchemy import or_ as _or
        q = db.query(Datasource).filter(_or(Datasource.user_id == "dev_user", Datasource.user_id.is_(None)))
        rows = q.order_by(Datasource.created_at.desc()).all()
        return [DatasourceOut.model_validate(r) for r in rows]
    owned = db.query(Datasource).filter(Datasource.user_id == uid)
    share_ids = [s.datasource_id for s in db.query(DatasourceShare).filter(DatasourceShare.user_id == uid).all()]
    shared = db.query(Datasource).filter(Datasource.id.in_(share_ids)) if share_ids else []
    owned_rows = owned.all()
    shared_rows = (shared.all() if share_ids else [])
    seen: set[str] = set()
    all_rows: list[Datasource] = []
    for r in sorted(owned_rows + shared_rows, key=lambda x: x.created_at, reverse=True):
        if r.id in seen:
            continue
        seen.add(r.id)
        all_rows.append(r)
    return [DatasourceOut.model_validate(r) for r in all_rows]


@router.get("/{ds_id}", response_model=DatasourceDetailOut)
def get_ds(ds_id: str, actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    ds: Datasource | None = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        t = str(ds.type or '').lower()
        is_public_duck = (t == 'duckdb') and not bool(ds.connection_encrypted)
        if (ds.user_id is not None and ds.user_id != actor) and (not is_public_duck):
            share = db.query(DatasourceShare).filter(DatasourceShare.datasource_id == ds.id, DatasourceShare.user_id == actor).first()
            if not share:
                raise HTTPException(status_code=403, detail="Forbidden")
    # Decrypt connection
    conn = None
    if ds.connection_encrypted:
        try:
            conn = decrypt_text(ds.connection_encrypted)
        except Exception:
            conn = None
    # Parse options
    try:
        opts = json.loads(ds.options_json or "{}")
    except Exception:
        opts = {}
    return DatasourceDetailOut(
        id=ds.id,
        name=ds.name,
        type=ds.type,
        created_at=ds.created_at,
        active=bool(getattr(ds, "active", True)),
        connectionUri=conn,
        options=opts,
    )


# Update datasource (edit dialog)
@router.patch("/{ds_id}", response_model=DatasourceOut)
def patch_ds(ds_id: str, payload: DatasourceUpdate, db: Session = Depends(get_db)):
    ds: Datasource | None = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    # Apply partial updates
    if payload.name is not None:
        ds.name = payload.name
    if payload.type is not None:
        ds.type = payload.type
    if payload.connectionUri is not None:
        ds.connection_encrypted = encrypt_text(payload.connectionUri) if payload.connectionUri else None
    if payload.options is not None:
        try:
            ds.options_json = json.dumps(payload.options)
        except Exception:
            ds.options_json = json.dumps({})
    if payload.active is not None:
        try:
            ds.active = bool(payload.active)
        except Exception:
            pass
    db.add(ds)
    db.commit()
    db.refresh(ds)
    return DatasourceOut.model_validate(ds)


@router.get("/{ds_id}/shares", response_model=list[DatasourceShareOut])
def list_ds_shares(ds_id: str, actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    shares = db.query(DatasourceShare).filter(DatasourceShare.datasource_id == ds_id).all()
    user_ids = [s.user_id for s in shares]
    users = db.query(User).filter(User.id.in_(user_ids)).all() if user_ids else []
    umap = {u.id: u for u in users}
    out: list[DatasourceShareOut] = []
    for s in shares:
        u = umap.get(s.user_id)
        out.append(DatasourceShareOut(userId=s.user_id, name=(u.name if u else None), email=(u.email if u else None), permission=s.permission, createdAt=s.created_at))
    return out


@router.post("/{ds_id}/shares", response_model=DatasourceShareOut)
def add_ds_share(ds_id: str, payload: DatasourceShareAddRequest, actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    tgt = db.get(User, str(payload.userId).strip())
    if not tgt:
        raise HTTPException(status_code=404, detail="User not found")
    existing = db.query(DatasourceShare).filter(DatasourceShare.datasource_id == ds_id, DatasourceShare.user_id == tgt.id).first()
    if existing:
        existing.permission = (payload.permission or existing.permission)
        db.add(existing)
        db.commit()
        db.refresh(existing)
        return DatasourceShareOut(userId=existing.user_id, name=tgt.name, email=tgt.email, permission=existing.permission, createdAt=existing.created_at)
    from uuid import uuid4
    s = DatasourceShare(id=str(uuid4()), datasource_id=ds_id, user_id=tgt.id, permission=(payload.permission or "ro"))
    db.add(s)
    db.commit()
    db.refresh(s)
    return DatasourceShareOut(userId=s.user_id, name=tgt.name, email=tgt.email, permission=s.permission, createdAt=s.created_at)


@router.delete("/{ds_id}/shares/{user_id}")
def remove_ds_share(ds_id: str, user_id: str, actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    db.query(DatasourceShare).filter(DatasourceShare.datasource_id == ds_id, DatasourceShare.user_id == str(user_id).strip()).delete()
    db.commit()
    return {"ok": True}


@router.put("/{ds_id}", response_model=DatasourceOut)
def put_ds(ds_id: str, payload: DatasourceUpdate, db: Session = Depends(get_db)):
    # Same semantics as PATCH for now
    return patch_ds(ds_id, payload, db)


@router.post("/{ds_id}/engine/dispose")
def dispose_ds_engine(ds_id: str, db: Session = Depends(get_db)):
    ds: Datasource | None = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    if not ds.connection_encrypted:
        ok = dispose_duck_engine()
        return {"disposed": bool(ok), "target": "duckdb"}
    try:
        dsn = decrypt_text(ds.connection_encrypted)
    except Exception:
        dsn = None
    if not dsn:
        return {"disposed": False, "target": "external", "message": "invalid connection"}
    ok = dispose_engine_by_key(dsn)
    return {"disposed": bool(ok), "target": "external"}


@router.post("/engines/dispose-all")
def dispose_all_cached_engines():
    count = dispose_all_engines()
    return {"disposed": int(count)}


@router.post("/{ds_id}/activate", response_model=DatasourceOut)
def activate_ds(ds_id: str, actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    ds: Datasource | None = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    try:
        ds.active = True
    except Exception:
        # older DBs may not have the column yet
        pass
    db.add(ds)
    db.commit()
    db.refresh(ds)
    return DatasourceOut.model_validate(ds)


@router.post("/{ds_id}/deactivate", response_model=DatasourceOut)
def deactivate_ds(ds_id: str, actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    ds: Datasource | None = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    # Only owner or admin can toggle active
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    try:
        ds.active = False
    except Exception:
        pass
    db.add(ds)
    db.commit()
    db.refresh(ds)
    return DatasourceOut.model_validate(ds)


# --- Sync tasks management ---
@router.get("/{ds_id}/sync-tasks", response_model=list[SyncTaskOut])
def list_sync_tasks(ds_id: str, actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    
    # For DuckDB datasources, find tasks by destination path
    is_duckdb = str(ds.type or '').lower().startswith('duckdb')
    
    if is_duckdb:
        # Get DuckDB path from connection_encrypted or use active path
        duck_path = None
        if ds.connection_encrypted:
            try:
                dsn = decrypt_text(ds.connection_encrypted)
                if dsn:
                    duck_path = dsn.replace('duckdb:///', '')
            except Exception:
                pass
        if not duck_path:
            duck_path = get_active_duck_path()
        
        # Normalize path for comparison (handle Windows slashes and casing)
        norm_duck_path = None
        if duck_path:
            norm_duck_path = os.path.normpath(duck_path).lower()

        all_tasks = db.query(SyncTask).order_by(SyncTask.created_at.asc()).all()
        tasks = []
        for t in all_tasks:
            st = db.query(SyncState).filter(SyncState.task_id == t.id).first()
            if st and st.last_duck_path and norm_duck_path:
                # Normalize stored path too
                st_path = os.path.normpath(st.last_duck_path).lower()
                if norm_duck_path in st_path or st_path in norm_duck_path:
                    tasks.append(t)
    else:
        tasks = db.query(SyncTask).filter(SyncTask.datasource_id == ds_id).order_by(SyncTask.created_at.asc()).all()
    
    return [_task_to_out(db, t) for t in tasks]


@router.post("/{ds_id}/sync-tasks", response_model=SyncTaskOut)
def create_sync_task(ds_id: str, payload: SyncTaskCreate, actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    # Normalize
    mode = (payload.mode or "").lower()
    if mode not in {"sequence", "snapshot"}:
        raise HTTPException(status_code=400, detail="mode must be 'sequence' or 'snapshot'")
    # For API datasources, sequence watermark comes from options.api; allow missing sequenceColumn
    if mode == "sequence" and not payload.sequenceColumn:
        ds_for_check = db.get(Datasource, ds_id)
        if not ds_for_check or (str(ds_for_check.type or '').lower() != 'api'):
            raise HTTPException(status_code=400, detail="sequenceColumn is required for sequence mode")
    dest = payload.destTableName.strip()
    if not dest:
        raise HTTPException(status_code=400, detail="destTableName is required")
    group_key = _group_key_for(ds_id, payload.sourceSchema, payload.sourceTable, dest)

    t = SyncTask(
        id=str(uuid4()),
        datasource_id=ds_id,
        source_schema=(payload.sourceSchema or None),
        source_table=payload.sourceTable,
        dest_table_name=dest,
        mode=mode,
        sequence_column=(payload.sequenceColumn or None),
        schedule_cron=(payload.scheduleCron or None),
        enabled=bool(payload.enabled),
        group_key=group_key,
        batch_size=(payload.batchSize or 10000),
        custom_query=(payload.customQuery.strip() if payload.customQuery and payload.customQuery.strip() else None),
    )
    t.pk_columns = payload.pkColumns or []
    t.select_columns = payload.selectColumns or []
    db.add(t)
    # Ensure a state row exists
    st = SyncState(id=str(uuid4()), task_id=t.id, in_progress=False)
    db.add(st)
    db.commit()
    db.refresh(t)
    return _task_to_out(db, t)


@router.get("/{ds_id}/sync/status", response_model=list[SyncTaskOut])
def get_sync_status(ds_id: str, actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    # Allow owner or admin to view
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    
    # For DuckDB datasources, find tasks by destination path
    is_duckdb = str(ds.type or '').lower().startswith('duckdb')
    
    if is_duckdb:
        # Get DuckDB path from connection_encrypted or use active path
        duck_path = None
        if ds.connection_encrypted:
            try:
                dsn = decrypt_text(ds.connection_encrypted)
                if dsn:
                    duck_path = dsn.replace('duckdb:///', '')
            except Exception:
                pass
        if not duck_path:
            duck_path = get_active_duck_path()
        
        # Normalize path for comparison (handle Windows slashes and casing)
        norm_duck_path = None
        if duck_path:
            norm_duck_path = os.path.normpath(duck_path).lower()

        print(f"[DEBUG] get_sync_status ds_id={ds_id} duck_path={duck_path} norm={norm_duck_path}", flush=True)

        all_tasks = db.query(SyncTask).all()
        tasks = []
        seen_ids = set()

        for t in all_tasks:
            # Match by ID directly (Primary)
            if str(t.datasource_id) == str(ds_id):
                tasks.append(t)
                seen_ids.add(t.id)
                continue

            # Match by Path (Secondary - for shared DBs)
            st = db.query(SyncState).filter(SyncState.task_id == t.id).first()
            if st and st.last_duck_path and norm_duck_path:
                st_path = os.path.normpath(st.last_duck_path).lower()
                # print(f"[DEBUG] Check task {t.id} st_path={st_path}", flush=True)
                if (norm_duck_path in st_path or st_path in norm_duck_path):
                    if t.id not in seen_ids:
                        tasks.append(t)
                        seen_ids.add(t.id)
    else:
        tasks = db.query(SyncTask).filter(SyncTask.datasource_id == ds_id).all()
    
    return [_task_to_out(db, t) for t in tasks]


@router.post("/{ds_id}/sync/run")
def run_sync_now(
    ds_id: str,
    response: Response,
    taskId: str | None = Query(default=None),
    execute: bool = Query(default=False),
    actorId: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    # Enforce permission: owner or admin
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    # Respect active flag
    try:
        if hasattr(ds, 'active') and not bool(ds.active):
            return {"ok": False, "message": "Datasource is inactive; activate it to run sync."}
    except Exception:
        pass
    # Enforce blackout windows at datasource level
    if _is_blackout_now(ds):
        return {"ok": False, "message": "Blackout period active; sync is not allowed now"}
    # Enforce max concurrent syncs per datasource
    maxc = _max_concurrent(ds)
    running = (
        db.query(SyncState)
        .join(SyncTask, SyncTask.id == SyncState.task_id)
        .filter(SyncTask.datasource_id == ds_id, SyncState.in_progress == True)
        .count()
    )
    if running >= maxc:
        return {"ok": False, "message": f"Max concurrent syncs reached ({running}/{maxc}). Try later."}
    duck_engine = get_duckdb_engine()
    # Resolve the current active DuckDB path at start of this run
    curr_duck_path = get_active_duck_path()
    source_engine = None
    is_api = str(ds.type or '').lower() == 'api'
    try:
        _log.info("run_sync_now start: ds_id=%s is_api=%s execute=%s taskId=%s", ds_id, is_api, execute, taskId)
        if DEBUG:
            print(f"[ds] run_sync_now start ds_id={ds_id} is_api={is_api} execute={execute} taskId={taskId}", flush=True)
    except Exception:
        pass
    if not is_api:
        if not ds.connection_encrypted:
            raise HTTPException(status_code=400, detail="Datasource has no connection URI for sync")
        dsn = decrypt_text(ds.connection_encrypted)
        if not dsn:
            raise HTTPException(status_code=400, detail="Invalid connection secret")
        source_engine = get_engine_from_dsn(dsn)

    if taskId:
        tasks = db.query(SyncTask).filter(SyncTask.id == taskId, SyncTask.datasource_id == ds_id).all()
    else:
        tasks = db.query(SyncTask).filter(SyncTask.datasource_id == ds_id, SyncTask.enabled == True).all()
    if not tasks:
        return {"ok": True, "message": "No tasks to run", "count": 0}

    # If not executing inline, enqueue and return 202 immediately
    if not execute:
        try:
            from ..scheduler import ensure_scheduler_started, run_task_job
            from apscheduler.triggers.date import DateTrigger
            from datetime import datetime as _dt
        except Exception:
            # Fallback: if scheduler unavailable, run inline
            execute = True
        if not execute:
            sched = ensure_scheduler_started()
            tasks_sorted = sorted(tasks, key=lambda t: 0 if t.mode == "snapshot" else 1)
            enq = 0
            for t in tasks_sorted:
                try:
                    # Use local now (not utcnow) for DateTrigger to avoid timezone misfire by ~3h
                    sched.add_job(
                        func=run_task_job,
                        trigger=DateTrigger(run_date=_dt.now()),
                        kwargs={"ds_id": ds_id, "task_id": t.id},
                        max_instances=1,
                        coalesce=True,
                        misfire_grace_time=60,
                    )
                    enq += 1
                except Exception:
                    continue
            try:
                response.status_code = 202
            except Exception:
                pass
            return {"ok": True, "message": "enqueued", "count": enq}

    # Run snapshots first, then sequences to avoid overlap in a single call
    tasks_sorted = sorted(tasks, key=lambda t: 0 if t.mode == "snapshot" else 1)
    print(f"[ABORT] Execute path: found {len(tasks_sorted)} tasks to sync", flush=True)
    results: list[dict] = []

    # Acquire locks per group key to ensure idempotent runs across processes
    acquired_keys: list[str] = []
    busy_keys: list[str] = []
    print(f"[ABORT] About to acquire locks for {len(tasks_sorted)} tasks", flush=True)
    try:
        uniq_groups = sorted({t.group_key for t in tasks_sorted})
        
        # When execute=True (force run), clear any existing locks for these groups
        if execute:
            for gk in uniq_groups:
                try:
                    deleted = db.query(SyncLock).filter(SyncLock.group_key == gk).delete()
                    if deleted:
                        print(f"[ABORT] Force-cleared existing lock for group: {gk}", flush=True)
                    db.commit()
                except Exception as clear_err:
                    print(f"[ABORT] Failed to clear lock for group {gk}: {clear_err}", flush=True)
                    try:
                        db.rollback()
                    except Exception:
                        pass
        
        for gk in uniq_groups:
            try:
                existed = db.query(SyncLock).filter(SyncLock.group_key == gk).first()
                if existed:
                    busy_keys.append(gk)
                    continue
                lk = SyncLock(group_key=gk)
                db.add(lk); db.commit()
                acquired_keys.append(gk)
            except Exception:
                # If we fail to create, consider it busy
                busy_keys.append(gk)
        if busy_keys:
            print(f"[ABORT] Cannot sync - groups locked: {busy_keys}", flush=True)
            try: counter_inc("sync_lock_busy_total")
            except Exception: pass
            # Release any previously acquired locks before returning
            for gk in acquired_keys:
                try: db.query(SyncLock).filter(SyncLock.group_key == gk).delete(); db.commit()
                except Exception: pass
            return {"ok": False, "message": f"Groups locked: {', '.join(busy_keys)}"}
        else:
            print(f"[ABORT] Successfully acquired locks for groups: {acquired_keys}", flush=True)
            try: counter_inc("sync_lock_acquired_total")
            except Exception: pass
    except Exception as lock_err:
        # Best-effort; proceed without hard lock if lock table unavailable
        print(f"[ABORT] Lock acquisition failed (will proceed anyway): {lock_err}", flush=True)
        pass

    # Parse options once
    try:
        ds_opts = json.loads(ds.options_json or '{}')
    except Exception:
        ds_opts = {}
    api_opts = (ds_opts.get('api') or {}) if is_api else None
    try:
        _log.info("options parsed: api_opts=%s keys=%s", bool(api_opts), (list(api_opts.keys()) if api_opts else []))
        if DEBUG:
            print(f"[ds] api_opts present={bool(api_opts)} keys={(list(api_opts.keys()) if api_opts else [])}", flush=True)
    except Exception:
        pass
    
    # Helper to check abort flag - uses separate session to avoid affecting main session cache
    def _check_abort(state_id: str) -> bool:
        try:
            # Use a separate session to avoid expiring the main session's objects
            check_db = SessionLocal()
            try:
                fresh_state = check_db.query(SyncState).filter(SyncState.id == state_id).first()
                abort_flag = bool(getattr(fresh_state, 'cancel_requested', False)) if fresh_state else False
                print(f"[ABORT_CHECK] state_id={state_id}, cancel_requested={abort_flag}, in_progress={getattr(fresh_state, 'in_progress', None) if fresh_state else None}", flush=True)
                return abort_flag
            finally:
                check_db.close()
        except Exception as e:
            print(f"[ABORT_CHECK] Error checking abort flag for state_id={state_id}: {e}", flush=True)
            return False

    print(f"[ABORT] Processing {len(tasks_sorted)} tasks for sync", flush=True)
    for t in tasks_sorted:
        print(f"[ABORT] Processing task: task_id={t.id}, mode={t.mode}, dest_table={t.dest_table_name}", flush=True)
        st = db.query(SyncState).filter(SyncState.task_id == t.id).first()
        if not st:
            print(f"[ABORT] Creating new SyncState for task_id={t.id}", flush=True)
            st = SyncState(id=str(uuid4()), task_id=t.id, in_progress=False)
            # For sequence tasks, initialize watermark from existing destination data if available
            if t.mode == "sequence" and t.sequence_column:
                try:
                    dest_name = t.dest_table_name
                    if str(os.getenv("DUCKDB_USER_SCOPED_TABLES", "0")).strip().lower() in ("1", "true", "yes", "on"):
                        owner = (actorId or ds.user_id or "dev_user").strip() or "dev_user"
                        safe = re.sub(r"[^A-Za-z0-9_]", "_", owner)
                        dest_name = f"{safe}__{t.dest_table_name}"
                    # Check if destination table exists and has data
                    if _duckdb is not None:
                        con = open_duck_native(curr_duck_path)
                        try:
                            def _q_duck(name: str) -> str:
                                return '"' + str(name).replace('"', '""') + '"'
                            # Check if table exists
                            check_sql = f"SELECT COUNT(*) FROM information_schema.tables WHERE table_name = '{dest_name}'"
                            table_exists = con.execute(check_sql).fetchone()[0] > 0
                            if table_exists:
                                # Get MAX(sequence_column) from existing data
                                max_sql = f"SELECT MAX({_q_duck(t.sequence_column)}) FROM {_q_duck(dest_name)}"
                                row = con.execute(max_sql).fetchone()
                                max_val = (row[0] if row else None)
                                if max_val is not None:
                                    st.last_sequence_value = int(max_val)
                                    print(f"[SEQUENCE] Initialized watermark for task_id={t.id} from existing data: {max_val}", flush=True)
                        except Exception as e:
                            print(f"[SEQUENCE] Could not initialize watermark from existing data: {e}", flush=True)
                            # Default to 0 for new sequence tasks without existing data
                            st.last_sequence_value = 0
                except Exception as e:
                    print(f"[SEQUENCE] Error initializing watermark for task_id={t.id}: {e}", flush=True)
                    st.last_sequence_value = 0
            db.add(st)
            db.commit()
        # Auto-reset watermark if the DuckDB file has changed since the last run.
        # When the active path switches, the destination table in the new file is empty,
        # so continuing from the old watermark would skip all historical rows.
        if t.mode == "sequence" and st.last_duck_path and curr_duck_path:
            prev_norm = os.path.normpath(st.last_duck_path).lower()
            curr_norm = os.path.normpath(curr_duck_path).lower()
            if prev_norm != curr_norm:
                print(f"[SEQUENCE] DuckDB path changed: {prev_norm} → {curr_norm}. Resetting watermark for task_id={t.id}", flush=True)
                st.last_sequence_value = None
                st.last_row_count = None

        # Mark in progress and clear any previous cancel flag
        print(f"[ABORT] Starting sync: state_id={st.id}, task_id={t.id}, setting in_progress=True, cancel_requested=False", flush=True)
        st.in_progress = True
        st.cancel_requested = False  # type: ignore[attr-defined]
        st.progress_current = 0
        st.progress_total = None
        st.progress_phase = 'fetch'  # type: ignore[attr-defined]
        st.started_at = datetime.now(timezone.utc)  # type: ignore[attr-defined]
        db.add(st)
        db.commit()
        print(f"[ABORT] Committed sync start for state_id={st.id}", flush=True)
        # Create a run log row
        run = SyncRun(id=str(uuid4()), task_id=t.id, datasource_id=ds_id, mode=t.mode)
        db.add(run)
        db.commit()
        try:
            try:
                _log.info("running task: id=%s mode=%s dest=%s", t.id, t.mode, t.dest_table_name if t.mode != "sequence" else t.dest_table_name)
                if DEBUG:
                    print(f"[ds] running task id={t.id} mode={t.mode} dest={t.dest_table_name}", flush=True)
            except Exception:
                pass
            if is_api:
                # API-driven ingestion; ignore source_engine
                cfg = dict(api_opts or {})
                # Allow per-task override in future via t.select_columns or source_table hints (not used yet)
                try:
                    _log.info("calling run_api_sync: endpoint=%s parse=%s query_keys=%s", (cfg.get('endpoint') or cfg.get('urlTemplate')), (cfg.get('parse') or cfg.get('format')), [ (q or {}).get('key') for q in (cfg.get('query') or []) ])
                    if DEBUG:
                        print(f"[ds] calling run_api_sync endpoint={(cfg.get('endpoint') or cfg.get('urlTemplate'))} parse={(cfg.get('parse') or cfg.get('format'))} qkeys={[ (q or {}).get('key') for q in (cfg.get('query') or []) ]}", flush=True)
                except Exception:
                    pass
                res = run_api_sync(
                    duck_engine=duck_engine,
                    options_api=cfg,
                    dest_table=t.dest_table_name,
                    mode=t.mode,
                )
                st.last_row_count = res.get("row_count")
                st.last_run_at = datetime.now(timezone.utc)
                st.error = None
                st.last_duck_path = curr_duck_path
                run.row_count = st.last_row_count
                run.finished_at = st.last_run_at
                try:
                    _log.info("run_api_sync returned: rows=%s window=%s..%s", st.last_row_count, res.get('windowStart'), res.get('windowEnd'))
                    if DEBUG:
                        print(f"[ds] run_api_sync returned rows={st.last_row_count} window={res.get('windowStart')}..{res.get('windowEnd')}", flush=True)
                except Exception:
                    pass
                results.append({"taskId": t.id, "mode": t.mode, "rowCount": st.last_row_count, "windowStart": res.get('windowStart'), "windowEnd": res.get('windowEnd')})
            elif t.mode == "sequence":
                # Optional user-scoped destination naming
                dest_name = t.dest_table_name
                try:
                    if str(os.getenv("DUCKDB_USER_SCOPED_TABLES", "0")).strip().lower() in ("1", "true", "yes", "on"):
                        owner = (actorId or ds.user_id or "dev_user").strip() or "dev_user"
                        safe = re.sub(r"[^A-Za-z0-9_]", "_", owner)
                        dest_name = f"{safe}__{t.dest_table_name}"
                except Exception:
                    pass
                res = run_sequence_sync(
                    source_engine,
                    duck_engine,
                    source_schema=t.source_schema,
                    source_table=t.source_table,
                    dest_table=dest_name,
                    sequence_column=t.sequence_column or "id",
                    pk_columns=t.pk_columns,
                    batch_size=int(t.batch_size or 10000),
                    last_sequence_value=st.last_sequence_value,
                    on_progress=lambda cur, tot: _update_progress(db, st.id, cur, tot),
                    select_columns=(t.select_columns or None),
                    should_abort=lambda: _check_abort(st.id),
                    on_phase=lambda ph: _set_phase(db, st.id, ph),
                    custom_query=(t.custom_query or None),
                )
                st.last_sequence_value = res.get("last_sequence_value")
                st.last_row_count = res.get("row_count")
                st.last_run_at = datetime.now(timezone.utc)
                st.error = ("aborted" if bool(res.get("aborted")) else None)
                st.last_duck_path = curr_duck_path
                run.row_count = st.last_row_count
                run.finished_at = st.last_run_at
                results.append({"taskId": t.id, "mode": t.mode, "rowCount": st.last_row_count, "lastSeq": st.last_sequence_value})
            else:  # snapshot
                dest_name = t.dest_table_name
                try:
                    if str(os.getenv("DUCKDB_USER_SCOPED_TABLES", "0")).strip().lower() in ("1", "true", "yes", "on"):
                        owner = (actorId or ds.user_id or "dev_user").strip() or "dev_user"
                        safe = re.sub(r"[^A-Za-z0-9_]", "_", owner)
                        dest_name = f"{safe}__{t.dest_table_name}"
                except Exception:
                    pass
                res = run_snapshot_sync(
                    source_engine,
                    duck_engine,
                    source_schema=t.source_schema,
                    source_table=t.source_table,
                    dest_table=dest_name,
                    batch_size=int(t.batch_size or 50000),
                    on_progress=lambda cur, tot: _update_progress(db, st.id, cur, tot),
                    select_columns=(t.select_columns or None),
                    should_abort=lambda: _check_abort(st.id),
                    on_phase=lambda ph: _set_phase(db, st.id, ph),
                    custom_query=(t.custom_query or None),
                )
                st.last_row_count = res.get("row_count")
                st.last_run_at = datetime.now(timezone.utc)
                st.error = ("aborted" if bool(res.get("aborted")) else None)
                st.last_duck_path = curr_duck_path
                run.row_count = st.last_row_count
                run.finished_at = st.last_run_at
                results.append({"taskId": t.id, "mode": t.mode, "rowCount": st.last_row_count})
                # After snapshot, set sequence tasks' watermark to MAX(sequence_column)
                seq_tasks = db.query(SyncTask).filter(SyncTask.group_key == t.group_key, SyncTask.mode == "sequence").all()
                try:
                    def _q_duck(name: str) -> str:
                        return '"' + str(name).replace('"', '""') + '"'
                    if _duckdb is not None:
                        con = open_duck_native(curr_duck_path)
                        try:
                            for sq in seq_tasks:
                                if not sq.sequence_column:
                                    continue
                                try:
                                    _dest = t.dest_table_name
                                    try:
                                        if str(os.getenv("DUCKDB_USER_SCOPED_TABLES", "0")).strip().lower() in ("1", "true", "yes", "on"):
                                            owner = (actorId or ds.user_id or "dev_user").strip() or "dev_user"
                                            safe = re.sub(r"[^A-Za-z0-9_]", "_", owner)
                                            _dest = f"{safe}__{t.dest_table_name}"
                                    except Exception:
                                        pass
                                    sql = f"SELECT MAX({_q_duck(sq.sequence_column)}) FROM {_q_duck(_dest)}"
                                    row = con.execute(sql).fetchone()
                                    mx = (row[0] if row else None)
                                    sst = db.query(SyncState).filter(SyncState.task_id == sq.id).first()
                                    if sst:
                                        sst.last_sequence_value = int(mx or 0)
                                        sst.last_run_at = st.last_run_at
                                        db.add(sst)
                                except Exception:
                                    pass
                        finally:
                            try:
                                con.close()
                            except Exception:
                                pass
                    else:
                        # Fallback to SQLAlchemy driver SQL with quoted identifiers
                        with duck_engine.connect() as conn:
                            for sq in seq_tasks:
                                if not sq.sequence_column:
                                    continue
                                try:
                                    _dest = t.dest_table_name
                                    try:
                                        if str(os.getenv("DUCKDB_USER_SCOPED_TABLES", "0")).strip().lower() in ("1", "true", "yes", "on"):
                                            owner = (ds.user_id or "dev_user").strip() or "dev_user"
                                            safe = re.sub(r"[^A-Za-z0-9_]", "_", owner)
                                            _dest = f"{safe}__{t.dest_table_name}"
                                    except Exception:
                                        pass
                                    sql = f"SELECT MAX({_q_duck(sq.sequence_column)}) FROM {_q_duck(_dest)}"
                                    mx = conn.exec_driver_sql(sql).scalar()
                                    sst = db.query(SyncState).filter(SyncState.task_id == sq.id).first()
                                    if sst:
                                        sst.last_sequence_value = int(mx or 0)
                                        sst.last_run_at = st.last_run_at
                                        db.add(sst)
                                except Exception:
                                    pass
                except Exception:
                    pass
        except Exception as e:
            st.error = str(e)
            run.error = st.error
            try:
                _log.error("task failed: %s", st.error)
            except Exception:
                pass
        finally:
            print(f"[ABORT] Finally block: Setting in_progress=False for state_id={st.id}, cancel_requested={getattr(st, 'cancel_requested', None)}", flush=True)
            st.in_progress = False
            try:
                st.progress_phase = None  # type: ignore[attr-defined]
                st.started_at = None  # type: ignore[attr-defined]
            except Exception:
                pass
            db.add(st)
            db.add(run)
            db.commit()
            print(f"[ABORT] Committed in_progress=False for state_id={st.id}", flush=True)

    # Release locks
    try:
        for gk in acquired_keys:
            try:
                db.query(SyncLock).filter(SyncLock.group_key == gk).delete()
                db.commit()
            except Exception:
                pass
    except Exception:
        pass
    return {"ok": True, "count": len(tasks_sorted), "results": results}


def _update_progress(db: Session, state_id: str, cur: int | None, tot: int | None) -> None:
    try:
        st = db.query(SyncState).filter(SyncState.id == state_id).first()
        if not st:
            return
        st.progress_current = int(cur or 0)
        st.progress_total = (int(tot) if tot is not None else None)
        db.add(st)
        db.commit()
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass


def _set_phase(db: Session, state_id: str, phase: str | None) -> None:
    try:
        st = db.query(SyncState).filter(SyncState.id == state_id).first()
        if not st:
            return
        st.progress_phase = (str(phase) if phase else None)
        db.add(st)
        db.commit()
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass

@router.get("/{ds_id}/sync/logs", response_model=list[SyncRunOut])
def list_sync_logs(ds_id: str, taskId: str | None = Query(default=None), limit: int = Query(default=50), actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    q = db.query(SyncRun).filter(SyncRun.datasource_id == ds_id)
    if taskId:
        q = q.filter(SyncRun.task_id == taskId)
    items = q.order_by(SyncRun.started_at.desc()).limit(int(max(1, min(limit, 500)))).all()
    return [
        SyncRunOut(
            id=it.id,
            taskId=it.task_id,
            datasourceId=it.datasource_id,
            mode=it.mode,
            startedAt=it.started_at,
            finishedAt=it.finished_at,
            rowCount=it.row_count,
            error=it.error,
        ) for it in items
    ]


@router.delete("/{ds_id}/sync/logs")
def clear_sync_logs(ds_id: str, taskId: str | None = Query(default=None), actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    q = db.query(SyncRun).filter(SyncRun.datasource_id == ds_id)
    if taskId:
        q = q.filter(SyncRun.task_id == taskId)
    deleted = 0
    try:
        deleted = int(q.delete(synchronize_session=False) or 0)
        db.commit()
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass
    return {"deleted": deleted}


@router.post("/{ds_id}/sync/abort")
def abort_sync(ds_id: str, taskId: str | None = Query(default=None), actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    from datetime import datetime, timedelta
    
    print(f"[ABORT] Endpoint called: ds_id={ds_id}, taskId={taskId}", flush=True)
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    
    q = db.query(SyncState).join(SyncTask, SyncTask.id == SyncState.task_id).filter(SyncTask.datasource_id == ds_id, SyncState.in_progress == True)
    if taskId:
        q = q.filter(SyncState.task_id == taskId)
    rows = q.all()
    print(f"[ABORT] Found {len(rows)} in-progress sync states to check", flush=True)
    
    cancel_requested = 0
    force_reset = 0
    stuck_threshold_minutes = 30  # Consider a sync stuck if in_progress for > 30 min without updates
    
    for st in rows:
        try:
            # Check if sync is actually stuck (no updates in last N minutes)
            is_stuck = False
            
            # Determine if stuck based on last_run_at timestamp on SyncState
            if st.last_run_at:
                last_update = st.last_run_at
                time_since_update = datetime.now(timezone.utc).replace(tzinfo=None) - last_update.replace(tzinfo=None)
                if time_since_update > timedelta(minutes=stuck_threshold_minutes):
                    is_stuck = True
                    print(f"[ABORT] Sync appears STUCK: state_id={st.id}, last_run_at={st.last_run_at}, minutes_since={time_since_update.total_seconds()/60:.1f}", flush=True)
                else:
                    print(f"[ABORT] Sync appears ACTIVE: state_id={st.id}, last_run_at={st.last_run_at}, minutes_since={time_since_update.total_seconds()/60:.1f}", flush=True)
            else:
                # No last_run_at means it never started properly - definitely stuck
                is_stuck = True
                print(f"[ABORT] Sync never started properly (no last_run_at): state_id={st.id}, forcing reset", flush=True)
            
            if is_stuck:
                # Force reset stuck sync
                print(f"[ABORT] FORCE RESET stuck sync: state_id={st.id}, task_id={st.task_id}", flush=True)
                st.in_progress = False
                st.cancel_requested = False  # type: ignore[attr-defined]
                st.progress_current = None
                st.progress_total = None
                st.progress_phase = None  # type: ignore[attr-defined]
                db.add(st)
                force_reset += 1
            else:
                # Active sync - just set cancel flag
                print(f"[ABORT] Setting cancel_requested=True for ACTIVE sync: state_id={st.id}, task_id={st.task_id}", flush=True)
                setattr(st, 'cancel_requested', True)
                db.add(st)
                cancel_requested += 1
                
        except Exception as e:
            print(f"[ABORT] Failed to process state_id={st.id}: {e}", flush=True)
            continue
    
    try:
        db.commit()
        print(f"[ABORT] Committed changes: {cancel_requested} active syncs flagged for cancellation, {force_reset} stuck syncs reset", flush=True)
    except Exception as e:
        print(f"[ABORT] Commit failed: {e}", flush=True)
        try: db.rollback()
        except Exception: pass
        raise HTTPException(status_code=500, detail=f"Failed to abort syncs: {e}")
    
    return {
        "ok": True, 
        "cancel_requested": cancel_requested,
        "force_reset": force_reset,
        "total_processed": cancel_requested + force_reset,
        "message": f"Flagged {cancel_requested} active sync(s) for cancellation, force reset {force_reset} stuck sync(s)"
    }


@router.post("/{ds_id}/sync/reset-stuck")
def reset_stuck_syncs(ds_id: str, actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    """
    Reset stuck sync states that are marked as in_progress but are not actually running.
    This is a recovery endpoint for when syncs crash or are killed without cleanup.
    """
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    
    # Find all stuck sync states (in_progress=True)
    stuck_states = (
        db.query(SyncState)
        .join(SyncTask, SyncTask.id == SyncState.task_id)
        .filter(SyncTask.datasource_id == ds_id, SyncState.in_progress == True)
        .all()
    )
    
    print(f"[RESET_STUCK] Found {len(stuck_states)} stuck sync states for datasource {ds_id}", flush=True)
    
    reset_count = 0
    for st in stuck_states:
        try:
            print(f"[RESET_STUCK] Resetting state_id={st.id}, task_id={st.task_id}", flush=True)
            st.in_progress = False
            st.cancel_requested = False  # type: ignore[attr-defined]
            st.progress_current = None
            st.progress_total = None
            st.progress_phase = None  # type: ignore[attr-defined]
            db.add(st)
            reset_count += 1
        except Exception as e:
            print(f"[RESET_STUCK] Failed to reset state_id={st.id}: {e}", flush=True)
            continue
    
    try:
        db.commit()
        print(f"[RESET_STUCK] Successfully reset {reset_count} stuck sync states", flush=True)
    except Exception as e:
        print(f"[RESET_STUCK] Commit failed: {e}", flush=True)
        try:
            db.rollback()
        except Exception:
            pass
        raise HTTPException(status_code=500, detail=f"Failed to reset stuck syncs: {e}")
    
    return {"ok": True, "reset_count": reset_count, "states_reset": [st.id for st in stuck_states]}


@router.patch("/{ds_id}/sync-tasks/{task_id}", response_model=SyncTaskOut)
def update_sync_task(ds_id: str, task_id: str, payload: SyncTaskCreate, actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    t = db.get(SyncTask, task_id)
    if not t or t.datasource_id != ds_id:
        raise HTTPException(status_code=404, detail="Task not found")
    ds = db.get(Datasource, ds_id)
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds and ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    mode = (payload.mode or t.mode).lower()
    if mode not in {"sequence", "snapshot"}:
        mode = t.mode
    t.source_schema = payload.sourceSchema if payload.sourceSchema is not None else t.source_schema
    t.source_table = payload.sourceTable or t.source_table
    t.dest_table_name = (payload.destTableName or t.dest_table_name).strip()
    t.mode = mode
    t.sequence_column = payload.sequenceColumn if payload.sequenceColumn is not None else t.sequence_column
    t.batch_size = int(payload.batchSize or t.batch_size or 10000)
    t.schedule_cron = payload.scheduleCron or None  # None = manual (no cron schedule)
    t.enabled = bool(payload.enabled) if payload.enabled is not None else t.enabled
    t.pk_columns = payload.pkColumns if payload.pkColumns is not None else t.pk_columns
    if payload.selectColumns is not None:
        t.select_columns = payload.selectColumns
    t.custom_query = (payload.customQuery.strip() if payload.customQuery and payload.customQuery.strip() else None)
    db.add(t)
    db.commit()
    db.refresh(t)
    # Sync scheduler immediately so job is added/removed without requiring manual refresh
    try:
        from ..scheduler import schedule_all_jobs
        schedule_all_jobs()
    except Exception:
        pass
    return _task_to_out(db, t)


@router.delete("/{ds_id}/sync-tasks/{task_id}", status_code=204)
def delete_sync_task(ds_id: str, task_id: str, actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    t = db.get(SyncTask, task_id)
    if not t or t.datasource_id != ds_id:
        return Response(status_code=204)
    ds = db.get(Datasource, ds_id)
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds and ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    # delete state and runs
    db.query(SyncState).filter(SyncState.task_id == task_id).delete()
    db.query(SyncRun).filter(SyncRun.task_id == task_id).delete()
    db.delete(t)
    db.commit()
    return Response(status_code=204)


@router.post("/{ds_id}/sync-tasks/{task_id}/flush")
def flush_sync_task(ds_id: str, task_id: str, actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    """Flush/reset a sync task's local DuckDB table without deleting the task itself.

    This drops the materialized table (including user-scoped variants) and resets the
    associated SyncState so that the next sync run will fully rebuild the table with
    freshly inferred types. Widgets and dashboards that reference the table by name
    continue to work after the table is rebuilt.
    """
    t = db.get(SyncTask, task_id)
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")

    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")

    # Determine whether this endpoint is being called in the context of a DuckDB
    # datasource (Data Model page) or the original remote/API datasource.
    is_duckdb = str(ds.type or "").lower().startswith("duckdb")

    # For non-DuckDB datasources, enforce that the task truly belongs to this
    # datasource. For DuckDB datasources, tasks may belong to remote sources but
    # share the same DuckDB file via last_duck_path, so we only require that the
    # task exists.
    if not is_duckdb and t.datasource_id != ds_id:
        raise HTTPException(status_code=404, detail="Task not found")

    # owner or admin
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")

    # Determine DuckDB file path
    duck_path = None
    if is_duckdb:
        # DuckDB datasource: derive path from its own connection URI
        if ds.connection_encrypted:
            try:
                dsn = decrypt_text(ds.connection_encrypted)
                if dsn:
                    duck_path = dsn.replace("duckdb:///", "")
            except Exception:
                duck_path = None
    if not duck_path:
        # Fallback to global active DuckDB path (used for remote->DuckDB syncs)
        try:
            duck_path = get_active_duck_path()
        except Exception:
            duck_path = settings.duckdb_path

    # Compute table names to drop: base dest and optional user-scoped variant
    dest = (t.dest_table_name or "").strip()
    if not dest:
        raise HTTPException(status_code=400, detail="Task has no destination table name")

    tables_to_drop: set[str] = set()
    tables_to_drop.add(dest)

    # User-scoped table naming mirrors run_sync_now logic
    try:
        if str(os.getenv("DUCKDB_USER_SCOPED_TABLES", "0")).strip().lower() in ("1", "true", "yes", "on"):
            owner = (actorId or ds.user_id or "dev_user").strip() or "dev_user"
            safe = re.sub(r"[^A-Za-z0-9_]", "_", owner)
            tables_to_drop.add(f"{safe}__{dest}")
    except Exception:
        pass

    dropped = 0
    if _duckdb is not None and duck_path:
        def _q_duck(name: str) -> str:
            return '"' + str(name).replace('"', '""') + '"'
        try:
            with open_duck_native(duck_path) as conn:
                for tbl in sorted(tables_to_drop):
                    try:
                        conn.execute(f"DROP TABLE IF EXISTS {_q_duck(tbl)}")
                        dropped += 1
                    except Exception:
                        # Best-effort per table
                        continue
        except Exception:
            # If we cannot open DuckDB, surface a 500 as this is a flush operation
            raise HTTPException(status_code=500, detail="Failed to open DuckDB for flush")

    # Reset sync state watermark and counters for this task
    st = db.query(SyncState).filter(SyncState.task_id == task_id).first()
    if st:
        st.last_sequence_value = None
        st.last_run_at = None
        st.last_row_count = None
        st.progress_current = None
        st.progress_total = None
        st.progress_phase = None  # type: ignore[attr-defined]
        st.error = None
        st.in_progress = False
        try:
            st.cancel_requested = False  # type: ignore[attr-defined]
        except Exception:
            pass
        st.last_duck_path = None
        db.add(st)
        db.commit()

    return {"ok": True, "dropped": int(dropped), "tables": sorted(tables_to_drop)}


# --- Local DuckDB stats for a datasource ---
@router.get("/{ds_id}/local/stats", response_model=LocalStatsResponse)
def local_stats(ds_id: str, db: Session = Depends(get_db)):
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    
    # For DuckDB datasources, find tasks by destination path, not source datasource_id
    is_duckdb = str(ds.type or '').lower().startswith('duckdb')
    
    if is_duckdb:
        # Get DuckDB path from connection_encrypted or use active path
        duck_path = None
        if ds.connection_encrypted:
            try:
                dsn = decrypt_text(ds.connection_encrypted)
                if dsn:
                    duck_path = dsn.replace('duckdb:///', '')
            except Exception:
                pass
        if not duck_path:
            duck_path = get_active_duck_path()
        
        # Find all tasks that write to this DuckDB (via their last_duck_path)
        all_tasks = db.query(SyncTask).all()
        all_states = {t.id: db.query(SyncState).filter(SyncState.task_id == t.id).first() for t in all_tasks}
        # Filter tasks that wrote to this specific DuckDB path
        tasks = []
        for t in all_tasks:
            st = all_states.get(t.id)
            if st and st.last_duck_path and duck_path and duck_path in st.last_duck_path:
                tasks.append(t)
        states_by_task = {t.id: all_states[t.id] for t in tasks}
    else:
        # For non-DuckDB datasources, use the original logic
        tasks = db.query(SyncTask).filter(SyncTask.datasource_id == ds_id).all()
        states_by_task: dict[str, SyncState | None] = { t.id: db.query(SyncState).filter(SyncState.task_id == t.id).first() for t in tasks }
    paths = []
    for t in tasks:
        st = states_by_task.get(t.id)
        p = (st.last_duck_path if st and st.last_duck_path else None)
        if not p:
            p = get_active_duck_path()
        paths.append(p)
    # Prefer a single representative path for header fields when consistent
    header_path = None
    try:
        uniq = sorted({p for p in paths if p})
        header_path = (uniq[0] if len(uniq) == 1 else get_active_duck_path())
    except Exception:
        header_path = get_active_duck_path()
    try:
        file_size = os.path.getsize(header_path) if header_path else 0
    except Exception:
        file_size = 0
    out: list[LocalTableStat] = []
    if _duckdb is None:
        raise HTTPException(status_code=500, detail="DuckDB native driver unavailable")
    try:
        def _q_duck(name: str) -> str:
            return '"' + str(name).replace('"', '""') + '"'
        
        # Get all tables from the DuckDB file directly
        path_to_query = header_path or get_active_duck_path()
        all_tables_in_duck: dict[str, dict] = {}  # table_name -> {row_count, ...}
        
        try:
            with open_duck_native(path_to_query) as conn:
                # Query information_schema to get all tables
                result = conn.execute("SELECT table_name FROM information_schema.tables WHERE table_schema = 'main'").fetchall()
                for row in result:
                    table_name = row[0]
                    try:
                        count_row = conn.execute(f"SELECT COUNT(*) FROM {_q_duck(table_name)}").fetchone()
                        row_count = int(count_row[0]) if count_row else 0
                    except Exception:
                        row_count = None
                    all_tables_in_duck[table_name] = {
                        'row_count': row_count,
                        'datasource_id': ds_id
                    }
        except Exception as e:
            # If we can't query DuckDB directly, fall back to task-based approach
            pass
        
        # Also include task info for tables that have sync tasks
        task_info: dict[str, dict] = {}  # table_name -> {lastSyncAt, sourceSchema, sourceTable}
        for t in tasks:
            st = states_by_task.get(t.id)
            task_info[t.dest_table_name] = {
                'lastSyncAt': (st.last_run_at if st else None),
                'sourceSchema': t.source_schema,
                'sourceTable': t.source_table,
            }
        
        # Combine: all tables from DuckDB + task info where available
        for table_name, table_data in all_tables_in_duck.items():
            task_data = task_info.get(table_name, {})
            out.append(
                LocalTableStat(
                    table=table_name,
                    rowCount=table_data.get('row_count'),
                    lastSyncAt=task_data.get('lastSyncAt'),
                    datasourceId=ds_id,
                    sourceSchema=task_data.get('sourceSchema'),
                    sourceTable=task_data.get('sourceTable'),
                )
            )
        
        return LocalStatsResponse(enginePath=header_path or get_active_duck_path(), fileSize=int(file_size), tables=out)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Local stats failed: {e}")


class _DropLocalTableRequest(BaseModel):
    table: str


@router.post("/{ds_id}/local/drop-table")
def drop_local_table(ds_id: str, payload: _DropLocalTableRequest, actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    # owner or admin
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    tbl = (payload.table or "").strip()
    if not tbl:
        raise HTTPException(status_code=400, detail="table is required")
    from ..db import open_duck_native
    dropped = 0
    # Quote for DuckDB to handle spaces/reserved words
    qtbl = '"' + tbl.replace('"', '""') + '"'
    try:
        with open_duck_native(settings.duckdb_path) as conn:
            try:
                conn.execute(f"DROP TABLE IF EXISTS {qtbl}")
                dropped = 1
            except Exception:
                dropped = 0
    except Exception:
        dropped = 0
    return {"ok": True, "dropped": dropped}


class _RenameLocalTableRequest(BaseModel):
    oldName: str
    newName: str


@router.post("/{ds_id}/local/rename-table")
def rename_local_table(ds_id: str, payload: _RenameLocalTableRequest, actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    # owner or admin
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    old_name = (payload.oldName or "").strip()
    new_name = (payload.newName or "").strip()
    if not old_name or not new_name:
        raise HTTPException(status_code=400, detail="oldName and newName are required")
    if old_name == new_name:
        return {"ok": True}
    from ..db import open_duck_native
    # Quote for DuckDB to handle spaces/reserved words
    q_old = '"' + old_name.replace('"', '""') + '"'
    q_new = '"' + new_name.replace('"', '""') + '"'
    try:
        with open_duck_native(settings.duckdb_path) as conn:
            conn.execute(f"ALTER TABLE {q_old} RENAME TO {q_new}")
        
        # Update table ID mapping to track renames
        try:
            opts = json.loads(ds.options_json or "{}")
        except Exception:
            opts = {}
        
        # Initialize table mappings if not present
        if "tableIdMappings" not in opts:
            opts["tableIdMappings"] = {}
        
        # Create stable table ID from datasource + original name
        table_id = f"{ds_id}__{old_name}"
        
        # Check if this table already has a mapping (was renamed before)
        # If so, keep the original ID, just update the current name
        found_id = None
        for tid, current_name in opts["tableIdMappings"].items():
            if current_name == old_name:
                found_id = tid
                break
        
        if found_id:
            # Update existing mapping
            opts["tableIdMappings"][found_id] = new_name
        else:
            # Create new mapping (first rename)
            opts["tableIdMappings"][table_id] = new_name
        
        # Save updated options
        ds.options_json = json.dumps(opts)
        db.commit()
        
        return {"ok": True}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Failed to rename table: {str(e)}")


def _to_json_safe(v):
    """Convert a DB value to something JSON-serialisable."""
    if v is None:
        return None
    if isinstance(v, (bool, int, str)):
        return v
    if isinstance(v, float):
        import math
        return None if math.isnan(v) or math.isinf(v) else v
    try:
        from decimal import Decimal
        if isinstance(v, Decimal):
            return float(v)
    except Exception:
        pass
    return str(v)


# --- Datasource-level transforms (Advanced SQL Mode) ---

@router.get("/{ds_id}/transforms", response_model=DatasourceTransforms)
def get_transforms(ds_id: str, db: Session = Depends(get_db)):
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    try:
        opts = json.loads(ds.options_json or "{}")
    except Exception:
        opts = {}
    cfg = (opts or {}).get("transforms") or {}
    # Normalize
    out = DatasourceTransforms(
        customColumns=cfg.get("customColumns", []),
        transforms=cfg.get("transforms", []),
        joins=cfg.get("joins", []),
        defaults=cfg.get("defaults"),
    )
    return out


@router.put("/{ds_id}/transforms", response_model=DatasourceTransforms)
def put_transforms(ds_id: str, payload: DatasourceTransforms, db: Session = Depends(get_db)):
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    try:
        opts = json.loads(ds.options_json or "{}")
    except Exception:
        opts = {}
    opts["transforms"] = payload.model_dump(by_alias=True)
    ds.options_json = json.dumps(opts)
    db.add(ds)
    db.commit()
    db.refresh(ds)
    return payload


@router.post("/{ds_id}/transforms/preview", response_model=PreviewResponse)
def preview_transforms(ds_id: str, payload: TransformsPreviewRequest, db: Session = Depends(get_db)):
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    # Infer dialect from datasource type string (keeps it simple for now)
    dialect = (ds.type or "").lower()
    source = payload.source or ""
    if not source:
        raise HTTPException(status_code=400, detail="source is required for preview")
    # Normalize payload items to plain dicts for sqlgen
    _cc_all = [ (c.model_dump(by_alias=True) if hasattr(c, 'model_dump') else c) for c in (payload.customColumns or []) ]
    _tr_all = [ (t.model_dump(by_alias=True) if hasattr(t, 'model_dump') else t) for t in (payload.transforms or []) ]
    _jn_all = [ (j.model_dump(by_alias=True) if hasattr(j, 'model_dump') else j) for j in (payload.joins or []) ]
    # Apply scope filtering for preview based on source and optional widgetId in context
    def _matches_table(scope_table: str, source_name: str) -> bool:
        def norm(s: str) -> str:
            s = (s or '').strip().strip('[]').strip('"').strip('`')
            parts = s.split('.')
            return parts[-1].lower()
        return norm(scope_table) == norm(source_name)
    widget_id = None
    try:
        widget_id = str(((payload.context or {}) or {}).get('widgetId') or '').strip()
    except Exception:
        widget_id = None
    def _filt(arr):
        out = []
        for it in (arr or []):
            sc = (it or {}).get('scope')
            if not sc:
                out.append(it); continue
            lvl = str((sc or {}).get('level') or '').lower()
            if lvl == 'datasource':
                out.append(it)
            elif lvl == 'table' and sc.get('table') and _matches_table(str(sc.get('table')), source):
                out.append(it)
            elif lvl == 'widget' and widget_id and str((sc or {}).get('widgetId') or '').strip() == widget_id:
                out.append(it)
        return out
    _cc = _filt(_cc_all)
    _tr = _filt(_tr_all)
    _jn = _filt(_jn_all)

    sql, cols, warns = build_sql(
        dialect=dialect,
        source=source,
        base_select=payload.select or ["*"],
        custom_columns=_cc,
        transforms=_tr,
        joins=_jn,
        defaults=payload.defaults or {},
        limit=payload.limit or 100,
    )
    # Execute against the datasource to return sample rows
    from sqlalchemy import text as _text
    from ..db import get_engine_from_dsn, get_duckdb_engine
    from ..security import decrypt_text
    if not ds.connection_encrypted:
        # Local DuckDB: prefer native to avoid engine mix
        from ..db import open_duck_native
        engine = None
    else:
        dsn = decrypt_text(ds.connection_encrypted)
        if not dsn:
            raise HTTPException(status_code=400, detail="Invalid connection secret")
        engine = get_engine_from_dsn(dsn)
    rows: list[list[Any]] = []
    cols_exec: list[str] | None = None
    try:
        if engine is None:
            # Native DuckDB path
            with open_duck_native(settings.duckdb_path) as conn:
                res = conn.execute(sql)
                cols_exec = [str(c[0]) for c in (res.description or [])]
                raw = res.fetchall()
                rows = [list(r) for r in raw]
        else:
            with engine.connect() as conn:
                res = conn.execute(_text(sql))
                cols_exec = [str(c) for c in res.keys()]
                raw = res.fetchall()
                rows = [list(r) for r in raw]
    except Exception as e:
        # Return SQL and warnings even if execution fails
        return PreviewResponse(sql=sql, columns=cols, rows=[], warnings=[*warns, f"exec error: {e}"])
    return PreviewResponse(sql=sql, columns=(cols_exec or cols), rows=rows, warnings=warns)


@router.get("/_local/schema", response_model=IntrospectResponse)
def introspect_local_schema():
    """Introspect the default local DuckDB without a datasource record."""
    if _duckdb is None:
        raise HTTPException(status_code=500, detail="DuckDB native driver unavailable")
    try:
        with open_duck_native(settings.duckdb_path) as conn:
            rows = conn.execute(
                """
                SELECT table_schema, table_name
                FROM information_schema.tables
                UNION
                SELECT table_schema, table_name
                FROM information_schema.views
                ORDER BY table_schema, table_name
                """
            ).fetchall()
            by_schema: dict[str, list[str]] = {}
            for sch, tbl in rows:
                by_schema.setdefault(str(sch), []).append(str(tbl))
            schemas: list[SchemaInfo] = []
            for sch, tbls in by_schema.items():
                tables: list[TableInfo] = []
                for tname in tbls:
                    cols_rows = conn.execute(
                        """
                        SELECT column_name, CAST(data_type AS VARCHAR) AS data_type
                        FROM information_schema.columns
                        WHERE table_schema = ? AND table_name = ?
                        ORDER BY ordinal_position
                        """,
                        (sch, tname),
                    ).fetchall()
                    cols = [ColumnInfo(name=str(cn), type=str(dt)) for (cn, dt) in cols_rows]
                    tables.append(TableInfo(name=tname, columns=cols))
                schemas.append(SchemaInfo(name=sch, tables=tables))
            return IntrospectResponse(schemas=schemas)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Introspection failed: {e}")


# Lightweight: tables-only (no columns) for faster table dropdowns
class _TablesSchema(BaseModel):
    name: str
    tables: list[str]

class TablesOnlyResponse(BaseModel):
    schemas: list[_TablesSchema]


@router.get("/_local/tables", response_model=TablesOnlyResponse)
def list_local_tables_only():
    if _duckdb is None:
        raise HTTPException(status_code=500, detail="DuckDB native driver unavailable")
    try:
        with open_duck_native(settings.duckdb_path) as conn:
            rows = conn.execute(
                """
                SELECT table_schema, table_name
                FROM information_schema.tables
                WHERE table_schema <> 'information_schema'
                  AND lower(table_name) NOT LIKE 'duckdb_%'
                  AND lower(table_name) NOT LIKE 'sqlite_%'
                  AND lower(table_name) NOT LIKE 'pragma_%'
                  AND lower(table_name) NOT IN ('sqlite_master','sqlite_temp_master','sqlite_schema','sqlite_temp_schema')
                ORDER BY table_schema, table_name
                """
            ).fetchall()
            by_schema: dict[str, list[str]] = {}
            for sch, tbl in rows:
                by_schema.setdefault(str(sch), []).append(str(tbl))
            out = [ _TablesSchema(name=sch, tables=tbls) for sch, tbls in by_schema.items() ]
            return TablesOnlyResponse(schemas=out)
    except Exception as e:
        mapped = _http_for_db_error(e)
        if mapped:
            raise mapped
        raise HTTPException(status_code=500, detail=f"List tables failed: {e}")


@router.get("/{ds_id}/schema", response_model=IntrospectResponse)
def introspect_schema(ds_id: str, db: Session = Depends(get_db)):
    print(f"[/schema] Called for datasource {ds_id}")
    try:
        ds: Datasource | None = db.get(Datasource, ds_id)
        if not ds:
            print(f"[/schema] ERROR: Datasource {ds_id} not found")
            raise HTTPException(status_code=404, detail="Datasource not found")
        print(f"[/schema] Found datasource: type={ds.type}, has_connection={bool(ds.connection_encrypted)}")
    except HTTPException:
        raise
    except Exception as e:
        print(f"[/schema] ERROR: Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Schema introspection failed: {e}")
    
    # Check if this is actually a remote datasource based on type
    ds_type_lower = str(ds.type or '').lower()
    is_remote = any(x in ds_type_lower for x in ['mssql', 'sqlserver', 'postgres', 'mysql', 'mariadb', 'oracle'])
    
    if not ds.connection_encrypted:
        if is_remote:
            print(f"[/schema] ERROR: Remote datasource ({ds.type}) has no connection string!")
            raise HTTPException(status_code=400, detail=f"Datasource is configured as {ds.type} but has no connection string. Please update the datasource configuration.")
        # Special-case: DuckDB datasource without a connection URI should introspect the local store
        try:
            if str(ds.type or '').lower().find('duckdb') != -1:
                if _duckdb is None:
                    raise HTTPException(status_code=500, detail="DuckDB native driver unavailable")
                try:
                    with open_duck_native(get_active_duck_path()) as conn:
                        rows = conn.execute(
                            """
                            SELECT table_schema, table_name
                            FROM information_schema.tables
                            WHERE table_schema <> 'information_schema'
                              AND lower(table_name) NOT LIKE 'duckdb_%'
                              AND lower(table_name) NOT LIKE 'sqlite_%'
                              AND lower(table_name) NOT LIKE 'pragma_%'
                              AND lower(table_name) NOT IN ('sqlite_master','sqlite_temp_master','sqlite_schema','sqlite_temp_schema')
                            UNION ALL
                            SELECT table_schema, table_name
                            FROM information_schema.views
                            WHERE table_schema <> 'information_schema'
                              AND lower(table_name) NOT LIKE 'duckdb_%'
                              AND lower(table_name) NOT LIKE 'sqlite_%'
                              AND lower(table_name) NOT LIKE 'pragma_%'
                              AND lower(table_name) NOT IN ('sqlite_master','sqlite_temp_master','sqlite_schema','sqlite_temp_schema')
                            ORDER BY table_schema, table_name
                            """
                        ).fetchall()
                        by_schema: dict[str, list[str]] = {}
                        for sch, tbl in rows:
                            by_schema.setdefault(str(sch), []).append(str(tbl))
                        schemas: list[SchemaInfo] = []
                        for sch, tbls in by_schema.items():
                            tables: list[TableInfo] = []
                            for tname in tbls:
                                cols_rows = conn.execute(
                                    """
                                    SELECT column_name, CAST(data_type AS VARCHAR) AS data_type
                                    FROM information_schema.columns
                                    WHERE table_schema = ? AND table_name = ?
                                    ORDER BY ordinal_position
                                    """,
                                    (sch, tname),
                                ).fetchall()
                                cols = [ColumnInfo(name=str(cn), type=str(dt)) for (cn, dt) in cols_rows]
                                tables.append(TableInfo(name=tname, columns=cols))
                            schemas.append(SchemaInfo(name=sch, tables=tables))
                        return IntrospectResponse(schemas=schemas)
                except HTTPException:
                    raise
                except Exception as e:
                    raise HTTPException(status_code=500, detail=f"Introspection failed: {e}")
        except HTTPException:
            raise
        # Non-DuckDB without connection: treat as error
        raise HTTPException(status_code=400, detail="Datasource has no connection URI")

    dsn = decrypt_text(ds.connection_encrypted)
    if not dsn:
        raise HTTPException(status_code=400, detail="Invalid connection secret")

    # If the source is DuckDB, use native driver directly and avoid SQLAlchemy entirely
    if (dsn or '').lower().startswith('duckdb:'):
        if _duckdb is None:
            raise HTTPException(status_code=500, detail="DuckDB native driver unavailable")
        # Extract filesystem path from DSN, handling absolute/relative and query params
        raw = dsn
        if raw.startswith('duckdb:////'):
            # Absolute path; preserve leading '/'
            path = '/' + raw[len('duckdb:////'):]
        elif raw.startswith('duckdb:///'):
            path = raw[len('duckdb:///'):]
        elif raw.startswith('duckdb://'):
            path = raw[len('duckdb://'):]
        else:
            path = raw[len('duckdb:'):]
        if '?' in path:
            path = path.split('?', 1)[0]
        path = unquote(path)
        if path in (':memory:', '/:memory:'):
            path = ':memory:'
        while path.startswith('//') and path != '://':
            path = path[1:]
        if not path:
            path = settings.duckdb_path
        # Normalize filesystem path
        if path != ':memory:':
            path = os.path.abspath(os.path.expanduser(path))
        try:
            with open_duck_native(path) as conn:
                rows = conn.execute(
                    """
                    SELECT table_schema, table_name FROM information_schema.tables
                    WHERE table_schema <> 'information_schema'
                      AND lower(table_name) NOT LIKE 'duckdb_%'
                      AND lower(table_name) NOT LIKE 'sqlite_%'
                      AND lower(table_name) NOT LIKE 'pragma_%'
                      AND lower(table_name) NOT IN ('sqlite_master','sqlite_temp_master','sqlite_schema','sqlite_temp_schema')
                    UNION ALL
                    SELECT table_schema, table_name FROM information_schema.views
                    WHERE table_schema <> 'information_schema'
                      AND lower(table_name) NOT LIKE 'duckdb_%'
                      AND lower(table_name) NOT LIKE 'sqlite_%'
                      AND lower(table_name) NOT LIKE 'pragma_%'
                      AND lower(table_name) NOT IN ('sqlite_master','sqlite_temp_master','sqlite_schema','sqlite_temp_schema')
                    ORDER BY table_schema, table_name
                    """
                ).fetchall()
                by_schema: dict[str, list[str]] = {}
                for sch, tbl in rows:
                    by_schema.setdefault(str(sch), []).append(str(tbl))
                schemas: list[SchemaInfo] = []
                for sch, tbls in by_schema.items():
                    tables: list[TableInfo] = []
                    for tname in tbls:
                        cols_rows = conn.execute(
                            "SELECT column_name, CAST(data_type AS VARCHAR) AS data_type FROM information_schema.columns WHERE table_schema = ? AND table_name = ? ORDER BY ordinal_position",
                            (sch, tname),
                        ).fetchall()
                        cols = [ColumnInfo(name=str(cn), type=str(dt)) for (cn, dt) in cols_rows]
                        tables.append(TableInfo(name=tname, columns=cols))
                    schemas.append(SchemaInfo(name=sch, tables=tables))
                return IntrospectResponse(schemas=schemas)
        except Exception as e:
            # Try falling back to the default local store path
            try:
                fallback = os.path.abspath(os.path.expanduser(settings.duckdb_path))
                with open_duck_native(fallback) as conn:
                    rows = conn.execute(
                        """
                        SELECT table_schema, table_name FROM information_schema.tables
                        WHERE table_schema <> 'information_schema'
                          AND lower(table_name) NOT LIKE 'duckdb_%'
                          AND lower(table_name) NOT LIKE 'sqlite_%'
                          AND lower(table_name) NOT LIKE 'pragma_%'
                          AND lower(table_name) NOT IN ('sqlite_master','sqlite_temp_master','sqlite_schema','sqlite_temp_schema')
                        UNION ALL
                        SELECT table_schema, table_name FROM information_schema.views
                        WHERE table_schema <> 'information_schema'
                          AND lower(table_name) NOT LIKE 'duckdb_%'
                          AND lower(table_name) NOT LIKE 'sqlite_%'
                          AND lower(table_name) NOT LIKE 'pragma_%'
                          AND lower(table_name) NOT IN ('sqlite_master','sqlite_temp_master','sqlite_schema','sqlite_temp_schema')
                        ORDER BY table_schema, table_name
                        """
                    ).fetchall()
                    by_schema: dict[str, list[str]] = {}
                    for sch, tbl in rows:
                        by_schema.setdefault(str(sch), []).append(str(tbl))
                    schemas: list[SchemaInfo] = []
                    for sch, tbls in by_schema.items():
                        tables: list[TableInfo] = []
                        for tname in tbls:
                            cols_rows = conn.execute(
                                "SELECT column_name, data_type FROM information_schema.columns WHERE table_schema = ? AND table_name = ? ORDER BY ordinal_position",
                                (sch, tname),
                            ).fetchall()
                            cols = [ColumnInfo(name=str(cn), type=str(dt)) for (cn, dt) in cols_rows]
                            tables.append(TableInfo(name=tname, columns=cols))
                        schemas.append(SchemaInfo(name=sch, tables=tables))
                    return IntrospectResponse(schemas=schemas)
            except Exception as e2:
                raise HTTPException(status_code=500, detail=f"DuckDB introspection failed for path '{path}': {e2}")
    # Non-duckdb: use SQLAlchemy
    try:
        engine = get_engine_from_dsn(dsn)
        insp = inspect(engine)
        schema_names = insp.get_schema_names()
        schemas: list[SchemaInfo] = []
        dialect_name = (engine.dialect.name or '').lower()
        is_mssql = 'mssql' in dialect_name or 'sqlserver' in dialect_name
        
        with engine.connect() as conn:
            for sch in schema_names:
                print(f"[/schema] Processing schema: {sch}")
                try:
                    tbls = insp.get_table_names(schema=sch)
                    print(f"[/schema]   Found {len(tbls)} tables")
                except Exception:
                    tbls = []
                # Try inspector first for views, then fall back to information_schema
                vws: list[str] = []
                try:
                    vws = [str(v) for v in insp.get_view_names(schema=sch)]
                    print(f"[/schema]   Found {len(vws)} views")
                except Exception:
                    vws = []
                if not vws:
                    try:
                        res = conn.execute(text("SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA = :s AND TABLE_TYPE = 'VIEW'"), {"s": sch})
                        vws = [str(r[0]) for r in res.fetchall()]
                        print(f"[/schema]   Found {len(vws)} views via INFORMATION_SCHEMA")
                    except Exception:
                        vws = []
                if not vws and is_mssql:
                    # SQL Server final fallback
                    try:
                        res = conn.execute(text("SELECT v.name FROM sys.views v JOIN sys.schemas s ON s.schema_id = v.schema_id WHERE s.name = :s"), {"s": sch})
                        vws = [str(r[0]) for r in res.fetchall()]
                        print(f"[/schema]   Found {len(vws)} views via sys.views")
                    except Exception:
                        vws = []
                names = list({*tbls, *vws})
                names.sort()
                
                # OPTIMIZATION: Batch-fetch ALL columns for this schema in one query
                print(f"[/schema]   Batch-fetching columns for {len(names)} tables...")
                columns_by_table: dict[str, list[ColumnInfo]] = {}
                try:
                    # Use INFORMATION_SCHEMA to get all columns for all tables in this schema at once
                    col_res = conn.execute(
                        text("SELECT TABLE_NAME, COLUMN_NAME, DATA_TYPE, ORDINAL_POSITION FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_SCHEMA = :s ORDER BY TABLE_NAME, ORDINAL_POSITION"),
                        {"s": sch}
                    )
                    for tbl_name, col_name, col_type, _ in col_res.fetchall():
                        tbl_name = str(tbl_name)
                        if tbl_name not in columns_by_table:
                            columns_by_table[tbl_name] = []
                        columns_by_table[tbl_name].append(ColumnInfo(name=str(col_name), type=str(col_type)))
                    print(f"[/schema]   Batch-fetched columns for {len(columns_by_table)} tables")
                except Exception as e:
                    print(f"[/schema]   WARNING: Batch column fetch failed: {e}, falling back to per-table queries")
                
                tables: list[TableInfo] = []
                for name in names:
                    # First, try to get from batch-fetched columns
                    cols = columns_by_table.get(name, [])
                    
                    # If not in batch (or batch failed), fall back to inspector
                    if not cols:
                        try:
                            cols = [ColumnInfo(name=str(c['name']), type=str(c.get('type'))) for c in insp.get_columns(name, schema=sch)]
                        except Exception:
                            cols = []
                    
                    tables.append(TableInfo(name=name, columns=cols))
                print(f"[/schema]   Created {len(tables)} table info objects")
                schemas.append(SchemaInfo(name=sch, tables=tables))
        return IntrospectResponse(schemas=schemas)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Introspection failed: {e}")


@router.get("/{ds_id}/tables", response_model=TablesOnlyResponse)
def list_tables_only(ds_id: str, db: Session = Depends(get_db)):
    print(f"[/tables] Called for datasource {ds_id}")
    try:
        ds: Datasource | None = db.get(Datasource, ds_id)
        if not ds:
            print(f"[/tables] ERROR: Datasource {ds_id} not found")
            raise HTTPException(status_code=404, detail="Datasource not found")
        print(f"[/tables] Found datasource: type={ds.type}, has_connection={bool(ds.connection_encrypted)}")
    except HTTPException:
        raise
    except Exception as e:
        print(f"[/tables] ERROR: Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"List tables failed: {e}")
    
    # Check if this is actually a remote datasource based on type
    ds_type_lower = str(ds.type or '').lower()
    is_remote = any(x in ds_type_lower for x in ['mssql', 'sqlserver', 'postgres', 'mysql', 'mariadb', 'oracle'])
    
    if not ds.connection_encrypted:
        if is_remote:
            print(f"[/tables] ERROR: Remote datasource ({ds.type}) has no connection string!")
            raise HTTPException(status_code=400, detail=f"Datasource is configured as {ds.type} but has no connection string. Please update the datasource configuration.")
        # Treat as DuckDB local — always use the actively tracked path, not the raw
        # settings default which may be stale/relative after a path switch.
        if _duckdb is None:
            raise HTTPException(status_code=500, detail="DuckDB native driver unavailable")
        try:
            with open_duck_native(get_active_duck_path()) as conn:
                rows = conn.execute(
                    """
                    SELECT table_schema, table_name FROM information_schema.tables
                    WHERE table_schema <> 'information_schema'
                      AND lower(table_name) NOT LIKE 'duckdb_%'
                      AND lower(table_name) NOT LIKE 'sqlite_%'
                      AND lower(table_name) NOT LIKE 'pragma_%'
                      AND lower(table_name) NOT IN ('sqlite_master','sqlite_temp_master','sqlite_schema','sqlite_temp_schema')
                    UNION
                    SELECT table_schema, table_name FROM information_schema.views
                    WHERE table_schema <> 'information_schema'
                      AND lower(table_name) NOT LIKE 'duckdb_%'
                      AND lower(table_name) NOT LIKE 'sqlite_%'
                      AND lower(table_name) NOT LIKE 'pragma_%'
                      AND lower(table_name) NOT IN ('sqlite_master','sqlite_temp_master','sqlite_schema','sqlite_temp_schema')
                    ORDER BY table_schema, table_name
                    """
                ).fetchall()
                by_schema: dict[str, set[str]] = {}
                for sch, tbl in rows:
                    by_schema.setdefault(str(sch), set()).add(str(tbl))
                out = [ _TablesSchema(name=sch, tables=sorted(list(tbls))) for sch, tbls in by_schema.items() ]
                return TablesOnlyResponse(schemas=out)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"List tables failed: {e}")
    # External datasource with connection string
    print(f"[/tables] Decrypting connection string...")
    try:
        dsn = decrypt_text(ds.connection_encrypted)
        print(f"[/tables] Decrypted successfully")
    except Exception as e:
        print(f"[/tables] ERROR: Decryption failed: {e}")
        dsn = None
    if not dsn:
        raise HTTPException(status_code=400, detail="Invalid connection secret")
    
    # Special handling for DuckDB - use native driver to avoid connection conflicts
    if (dsn or '').lower().startswith('duckdb:'):
        print(f"[/tables] Using native DuckDB driver...")
        if _duckdb is None:
            raise HTTPException(status_code=500, detail="DuckDB native driver unavailable")
        # Extract path from DSN
        raw = dsn
        if raw.startswith('duckdb:////'):
            path = '/' + raw[len('duckdb:////'):]
        elif raw.startswith('duckdb:///'):
            path = raw[len('duckdb:///'):]
        elif raw.startswith('duckdb://'):
            path = raw[len('duckdb://'):]
        else:
            path = raw[len('duckdb:'):]
        if '?' in path:
            path = path.split('?', 1)[0]
        from urllib.parse import unquote
        path = unquote(path)
        if path in (':memory:', '/:memory:'):
            path = ':memory:'
        while path.startswith('//') and path != '://':
            path = path[1:]
        if not path:
            path = settings.duckdb_path
        if path != ':memory:':
            import os
            path = os.path.abspath(os.path.expanduser(path))
        try:
            with open_duck_native(path) as conn:
                rows = conn.execute(
                    """
                    SELECT table_schema, table_name FROM information_schema.tables
                    WHERE table_schema <> 'information_schema'
                      AND lower(table_name) NOT LIKE 'duckdb_%'
                      AND lower(table_name) NOT LIKE 'sqlite_%'
                      AND lower(table_name) NOT LIKE 'pragma_%'
                      AND lower(table_name) NOT IN ('sqlite_master','sqlite_temp_master','sqlite_schema','sqlite_temp_schema')
                    UNION
                    SELECT table_schema, table_name FROM information_schema.views
                    WHERE table_schema <> 'information_schema'
                      AND lower(table_name) NOT LIKE 'duckdb_%'
                      AND lower(table_name) NOT LIKE 'sqlite_%'
                      AND lower(table_name) NOT LIKE 'pragma_%'
                      AND lower(table_name) NOT IN ('sqlite_master','sqlite_temp_master','sqlite_schema','sqlite_temp_schema')
                    ORDER BY table_schema, table_name
                    """
                ).fetchall()
                by_schema: dict[str, set[str]] = {}
                for sch, tbl in rows:
                    by_schema.setdefault(str(sch), set()).add(str(tbl))
                out = [ _TablesSchema(name=sch, tables=sorted(list(tbls))) for sch, tbls in by_schema.items() ]
                print(f"[/tables] Found {len(out)} DuckDB schemas")
                return TablesOnlyResponse(schemas=out)
        except Exception as e:
            print(f"[/tables] WARNING: DuckDB path '{path}' failed ({e}), falling back to active path")
            # DSN path may be stale (file deleted/moved) — fall back to the currently active store
            try:
                fallback = get_active_duck_path()
                if fallback != path:
                    with open_duck_native(fallback) as conn2:
                        rows = conn2.execute(
                            """
                            SELECT table_schema, table_name FROM information_schema.tables
                            WHERE table_schema <> 'information_schema'
                              AND lower(table_name) NOT LIKE 'duckdb_%'
                              AND lower(table_name) NOT LIKE 'sqlite_%'
                              AND lower(table_name) NOT LIKE 'pragma_%'
                              AND lower(table_name) NOT IN ('sqlite_master','sqlite_temp_master','sqlite_schema','sqlite_temp_schema')
                            UNION
                            SELECT table_schema, table_name FROM information_schema.views
                            WHERE table_schema <> 'information_schema'
                              AND lower(table_name) NOT LIKE 'duckdb_%'
                              AND lower(table_name) NOT LIKE 'sqlite_%'
                              AND lower(table_name) NOT LIKE 'pragma_%'
                              AND lower(table_name) NOT IN ('sqlite_master','sqlite_temp_master','sqlite_schema','sqlite_temp_schema')
                            ORDER BY table_schema, table_name
                            """
                        ).fetchall()
                        by_schema2: dict[str, set[str]] = {}
                        for sch, tbl in rows:
                            by_schema2.setdefault(str(sch), set()).add(str(tbl))
                        out2 = [_TablesSchema(name=sch, tables=sorted(list(tbls))) for sch, tbls in by_schema2.items()]
                        print(f"[/tables] Fallback found {len(out2)} schemas")
                        return TablesOnlyResponse(schemas=out2)
            except Exception as e2:
                print(f"[/tables] Fallback also failed: {e2}")
            raise HTTPException(status_code=500, detail=f"DuckDB tables query failed: {e}")
    
    # Non-DuckDB: use SQLAlchemy inspector
    try:
        print(f"[/tables] Creating engine...")
        engine = get_engine_from_dsn(dsn)
        print(f"[/tables] Getting inspector...")
        insp = inspect(engine)
        print(f"[/tables] Getting schema names...")
        schema_names = insp.get_schema_names()
        print(f"[/tables] Found {len(schema_names)} schemas")
        
        # Filter out SQL Server system schemas to speed up
        dialect_name = (engine.dialect.name or '').lower()
        if 'mssql' in dialect_name or 'sqlserver' in dialect_name:
            system_schemas = {'sys', 'INFORMATION_SCHEMA', 'guest', 'db_owner', 'db_accessadmin', 'db_securityadmin', 
                            'db_ddladmin', 'db_backupoperator', 'db_datareader', 'db_datawriter', 'db_denydatareader', 
                            'db_denydatawriter'}
            schema_names = [s for s in schema_names if s not in system_schemas]
            print(f"[/tables] Filtered to {len(schema_names)} user schemas")
        
        out: list[_TablesSchema] = []
        print(f"[/tables] Connecting to database...")
        with engine.connect() as conn:
            print(f"[/tables] Connected successfully")
            for sch in schema_names:
                print(f"[/tables] Processing schema: {sch}")
                try:
                    print(f"[/tables]   Getting tables...")
                    tbls = insp.get_table_names(schema=sch)
                    print(f"[/tables]   Found {len(tbls)} tables")
                except Exception as e:
                    print(f"[/tables]   Error getting tables: {e}")
                    tbls = []
                
                print(f"[/tables]   Getting views...")
                vws: list[str] = []
                try:
                    vws = [str(v) for v in insp.get_view_names(schema=sch)]
                    print(f"[/tables]   Found {len(vws)} views via inspector")
                except Exception as e:
                    print(f"[/tables]   Inspector failed, trying INFORMATION_SCHEMA: {e}")
                    # Only try SQL Server sys.views as fallback
                    try:
                        res = conn.execute(text("SELECT v.name FROM sys.views v JOIN sys.schemas s ON s.schema_id = v.schema_id WHERE s.name = :s"), {"s": sch})
                        vws = [str(r[0]) for r in res.fetchall()]
                        print(f"[/tables]   Found {len(vws)} views via sys.views")
                    except Exception as e2:
                        print(f"[/tables]   sys.views also failed: {e2}")
                        vws = []
                
                # Union tables and views; keep unique, sorted for stability
                combined = sorted(list({*(str(t) for t in tbls), *vws}))
                print(f"[/tables]   Total: {len(combined)} objects in schema {sch}")
                out.append(_TablesSchema(name=sch, tables=combined))
        
        print(f"[/tables] Finished processing {len(out)} schemas, returning response...")
        try:
            response = TablesOnlyResponse(schemas=out)
            print(f"[/tables] Response object created successfully")
            print(f"[/tables] Response has {sum(len(s.tables) for s in response.schemas)} total tables")
            return response
        except Exception as resp_err:
            print(f"[/tables] ERROR creating response: {resp_err}")
            import traceback
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Failed to create response: {resp_err}")
    except Exception as e:
        print(f"[/tables] ERROR in main block: {e}")
        import traceback
        traceback.print_exc()
        mapped = _http_for_db_error(e)
        if mapped:
            raise mapped
        raise HTTPException(status_code=500, detail=f"List tables failed: {e}")


@router.delete("/{ds_id}", status_code=204)
def delete_ds(ds_id: str, db: Session = Depends(get_db)):
    ds: Datasource | None = db.get(Datasource, ds_id)
    if ds:
        db.delete(ds)
        db.commit()
    # Idempotent: return 204 even if it wasn't found
    return Response(status_code=204)


# --- Export / Import: Datasources ---

def _to_export_item(ds: Datasource) -> DatasourceExportItem:
    # Decrypt connection
    conn = None
    if ds.connection_encrypted:
        try:
            conn = decrypt_text(ds.connection_encrypted)
        except Exception:
            conn = None
    try:
        opts = json.loads(ds.options_json or "{}")
    except Exception:
        opts = {}
    return DatasourceExportItem(
        id=ds.id,
        name=ds.name,
        type=ds.type,
        connectionUri=conn,
        options=opts,
        userId=ds.user_id,
        active=bool(getattr(ds, "active", True)),
        createdAt=ds.created_at,
    )


@router.get("/export", response_model=list[DatasourceExportItem])
def export_datasources(ids: list[str] | None = Query(default=None), includeSyncTasks: bool = Query(default=True), actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    # Permissions: admin can export any; non-admin can only export own datasources
    if _is_admin(db, actorId):
        q = db.query(Datasource)
    else:
        actor = (actorId or "").strip()
        if not actor:
            raise HTTPException(status_code=403, detail="Forbidden")
        q = db.query(Datasource).filter(Datasource.user_id == actor)
    if ids:
        q = q.filter(Datasource.id.in_(ids))
    rows = q.order_by(Datasource.created_at.desc()).all()
    out: list[DatasourceExportItem] = []
    for r in rows:
        item = _to_export_item(r)
        if includeSyncTasks:
            tasks = db.query(SyncTask).filter(SyncTask.datasource_id == r.id).order_by(SyncTask.created_at.asc()).all()
            st_items: list[dict] = []
            for t in tasks:
                st_items.append({
                    "id": t.id,
                    "datasourceId": t.datasource_id,
                    "sourceSchema": t.source_schema,
                    "sourceTable": t.source_table,
                    "destTableName": t.dest_table_name,
                    "mode": t.mode,
                    "pkColumns": t.pk_columns,
                    "selectColumns": t.select_columns,
                    "sequenceColumn": t.sequence_column,
                    "batchSize": t.batch_size,
                    "scheduleCron": t.schedule_cron,
                    "enabled": t.enabled,
                    "groupKey": t.group_key,
                    "createdAt": t.created_at,
                })
            item.syncTasks = st_items  # type: ignore[attr-defined]
        out.append(item)
    return out


@router.get("/{ds_id}/export", response_model=DatasourceExportItem)
def export_single_datasource(ds_id: str, includeSyncTasks: bool = Query(default=True), actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    if not _is_admin(db, actorId):
        actor = (actorId or "").strip()
        if ds.user_id and ds.user_id != actor:
            raise HTTPException(status_code=403, detail="Forbidden")
    item = _to_export_item(ds)
    if includeSyncTasks:
        tasks = db.query(SyncTask).filter(SyncTask.datasource_id == ds.id).order_by(SyncTask.created_at.asc()).all()
        st_items: list[dict] = []
        for t in tasks:
            st_items.append({
                "id": t.id,
                "datasourceId": t.datasource_id,
                "sourceSchema": t.source_schema,
                "sourceTable": t.source_table,
                "destTableName": t.dest_table_name,
                "mode": t.mode,
                "pkColumns": t.pk_columns,
                "selectColumns": t.select_columns,
                "sequenceColumn": t.sequence_column,
                "batchSize": t.batch_size,
                "scheduleCron": t.schedule_cron,
                "enabled": t.enabled,
                "groupKey": t.group_key,
                "createdAt": t.created_at,
            })
        item.syncTasks = st_items  # type: ignore[attr-defined]
    return item


@router.post("/import", response_model=DatasourceImportResponse)
def import_datasources(payload: DatasourceImportRequest, actorId: str | None = Query(default=None), db: Session = Depends(get_db)):
    if not payload or not isinstance(payload.items, list):
        raise HTTPException(status_code=400, detail="items array is required")
    created = 0
    updated = 0
    out: list[DatasourceOut] = []
    id_map: dict[str, str] = {}
    for it in payload.items:
        # Permission: admin may import for any user; otherwise force to actor
        target_user = (it.userId or actorId or "").strip()
        if not _is_admin(db, actorId):
            actor = (actorId or "").strip()
            if not actor:
                raise HTTPException(status_code=403, detail="Forbidden")
            # If item had userId different from actor, override to actor
            target_user = actor
        # Upsert by (name, user_id)
        existing = (
            db.query(Datasource)
            .filter(Datasource.name == it.name, Datasource.user_id == (target_user or None))
            .first()
        )
        enc = encrypt_text(it.connectionUri) if it.connectionUri else None
        if existing:
            existing.type = it.type
            existing.connection_encrypted = enc
            try:
                existing.active = bool(it.active) if it.active is not None else existing.active
            except Exception:
                pass
            try:
                existing.options_json = json.dumps(it.options or {})
            except Exception:
                existing.options_json = "{}"
            db.add(existing)
            db.commit()
            db.refresh(existing)
            updated += 1
            out.append(DatasourceOut.model_validate(existing))
            if it.id:
                try:
                    id_map[str(it.id)] = existing.id
                except Exception:
                    pass
        else:
            ds = create_datasource(
                db,
                NewDatasourceInput(
                    name=it.name,
                    type=it.type,
                    connection_encrypted=enc,
                    options=it.options,
                    user_id=target_user or None,
                ),
            )
            try:
                ds.active = bool(it.active) if it.active is not None else True
            except Exception:
                pass
            db.add(ds)
            db.commit()
            db.refresh(ds)
            created += 1
            out.append(DatasourceOut.model_validate(ds))
            if it.id:
                try:
                    id_map[str(it.id)] = ds.id
                except Exception:
                    pass
        # Import sync tasks for this datasource if provided
        try:
            if getattr(it, "syncTasks", None):
                for st in (it.syncTasks or []):  # type: ignore[attr-defined]
                    # Upsert by dest table name within this datasource
                    dest = (st.destTableName or "").strip()
                    if not dest:
                        continue
                    existing_task = (
                        db.query(SyncTask)
                        .filter(SyncTask.datasource_id == (existing.id if 'existing' in locals() and existing else ds.id), SyncTask.dest_table_name == dest)
                        .first()
                    )
                    tgt_ds_id = (existing.id if 'existing' in locals() and existing else ds.id)
                    mode = (st.mode or "snapshot").lower()
                    group_key = _group_key_for(tgt_ds_id, st.sourceSchema, st.sourceTable, dest)
                    if existing_task:
                        t = existing_task
                        t.source_schema = st.sourceSchema if st.sourceSchema is not None else t.source_schema
                        t.source_table = st.sourceTable or t.source_table
                        t.dest_table_name = dest
                        t.mode = mode
                        t.sequence_column = st.sequenceColumn if st.sequenceColumn is not None else t.sequence_column
                        t.batch_size = int(st.batchSize or t.batch_size or 10000)
                        t.schedule_cron = st.scheduleCron if st.scheduleCron is not None else t.schedule_cron
                        t.enabled = bool(st.enabled) if st.enabled is not None else t.enabled
                        t.pk_columns = (st.pkColumns or t.pk_columns)
                        t.select_columns = (st.selectColumns or t.select_columns)
                        t.group_key = group_key
                        db.add(t)
                    else:
                        t = SyncTask(
                            id=str(uuid4()),
                            datasource_id=tgt_ds_id,
                            source_schema=(st.sourceSchema or None),
                            source_table=st.sourceTable,
                            dest_table_name=dest,
                            mode=mode,
                            sequence_column=(st.sequenceColumn or None),
                            schedule_cron=(st.scheduleCron or None),
                            enabled=bool(st.enabled) if st.enabled is not None else True,
                            group_key=group_key,
                            batch_size=(st.batchSize or 10000),
                        )
                        t.pk_columns = st.pkColumns or []
                        t.select_columns = st.selectColumns or []
                        db.add(t)
                        # Ensure state row exists
                        st_row = SyncState(id=str(uuid4()), task_id=t.id, in_progress=False)
                        db.add(st_row)
                db.commit()
        except Exception:
            # Non-fatal; continue importing core datasources
            pass
    return DatasourceImportResponse(created=created, updated=updated, items=out, idMap=(id_map or None))


# ---------------------------------------------------------------------------
# Import Table — preview and commit (SQL query or file upload)
# ---------------------------------------------------------------------------

class _ImportSqlPreviewRequest(BaseModel):
    sql: str
    sourceDsId: str | None = None  # external source; omit to query local DuckDB
    limit: int = 100


class _ImportSqlCommitRequest(BaseModel):
    sql: str
    sourceDsId: str | None = None  # external source; omit to run SQL on local DuckDB
    tableName: str
    ifExists: str = "replace"   # "replace" | "append" | "fail"


def _duck_path_for_ds(ds: Datasource) -> str:
    """Return the DuckDB file path for a given datasource record."""
    if ds.connection_encrypted:
        try:
            dsn = decrypt_text(ds.connection_encrypted) or ""
            p = dsn.replace("duckdb:///", "").replace("duckdb://", "").strip()
            if p:
                return p
        except Exception:
            pass
    return get_active_duck_path()


def _q(name: str) -> str:
    """Quote an identifier for DuckDB."""
    return '"' + str(name).replace('"', '""') + '"'


@router.post("/{ds_id}/local/import-sql-preview")
def import_sql_preview(
    ds_id: str,
    payload: _ImportSqlPreviewRequest,
    db: Session = Depends(get_db),
):
    """Preview SQL: routes to external datasource when sourceDsId is provided, else queries local DuckDB."""
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    sql = (payload.sql or "").strip().rstrip(";")
    if not sql:
        raise HTTPException(status_code=400, detail="sql is required")
    limit = max(1, min(int(payload.limit or 100), 1000))
    # --- External datasource path ---
    if payload.sourceDsId:
        src_ds = db.get(Datasource, payload.sourceDsId)
        if not src_ds:
            raise HTTPException(status_code=404, detail="Source datasource not found")
        if not src_ds.connection_encrypted:
            raise HTTPException(status_code=400, detail="Source datasource has no connection URI")
        dsn = decrypt_text(src_ds.connection_encrypted)
        if not dsn:
            raise HTTPException(status_code=400, detail="Invalid source connection secret")
        try:
            source_engine = get_engine_from_dsn(dsn)
            with source_engine.connect() as conn:
                res = conn.execute(text(sql))
                columns = list(res.keys())
                rows_raw = res.fetchmany(limit)
                rows = [dict(zip(columns, [_to_json_safe(v) for v in row])) for row in rows_raw]
            return {"columns": columns, "rows": rows, "rowCount": len(rows)}
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Query failed: {e}")
    # --- Local DuckDB path ---
    if _duckdb is None:
        raise HTTPException(status_code=500, detail="DuckDB driver unavailable")
    duck_path = _duck_path_for_ds(ds)
    try:
        with open_duck_native(duck_path) as conn:
            res = conn.execute(f"SELECT * FROM ({sql}) __q LIMIT {limit}")
            columns = [str(c[0]) for c in (res.description or [])]
            raw = res.fetchall()
        rows = [dict(zip(columns, row)) for row in raw]
        rows = [{k: _to_json_safe(v) for k, v in r.items()} for r in rows]
        return {"columns": columns, "rows": rows, "rowCount": len(rows)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"SQL error: {e}")


@router.post("/{ds_id}/local/import-sql-commit")
def import_sql_commit(
    ds_id: str,
    payload: _ImportSqlCommitRequest,
    db: Session = Depends(get_db),
):
    """Create (or replace) a DuckDB table from SQL; routes to external source when sourceDsId provided."""
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    if _duckdb is None:
        raise HTTPException(status_code=500, detail="DuckDB driver unavailable")
    sql = (payload.sql or "").strip().rstrip(";")
    if not sql:
        raise HTTPException(status_code=400, detail="sql is required")
    table_name = re.sub(r"[^a-zA-Z0-9_]", "_", (payload.tableName or "").strip())
    if not table_name:
        raise HTTPException(status_code=400, detail="tableName is required")
    if_exists = str(payload.ifExists or "replace").lower()
    duck_path = _duck_path_for_ds(ds)
    qtbl = _q(table_name)
    # --- External datasource path: stream from source → insert into DuckDB ---
    if payload.sourceDsId:
        src_ds = db.get(Datasource, payload.sourceDsId)
        if not src_ds:
            raise HTTPException(status_code=404, detail="Source datasource not found")
        if not src_ds.connection_encrypted:
            raise HTTPException(status_code=400, detail="Source datasource has no connection URI")
        dsn = decrypt_text(src_ds.connection_encrypted)
        if not dsn:
            raise HTTPException(status_code=400, detail="Invalid source connection secret")
        try:
            source_engine = get_engine_from_dsn(dsn)
            total_rows = 0
            with source_engine.connect() as src:
                try:
                    stream_res = src.execution_options(stream_results=True).execute(text(sql))
                except Exception:
                    stream_res = src.execute(text(sql))
                columns = [c.strip() for c in stream_res.keys()]
                with open_duck_native(duck_path) as duck:
                    tbl_exists = duck.execute(
                        "SELECT COUNT(*) FROM information_schema.tables WHERE table_schema='main' AND table_name=?",
                        [table_name]
                    ).fetchone()[0] > 0
                    if tbl_exists:
                        if if_exists == "fail":
                            raise HTTPException(status_code=409, detail=f"Table '{table_name}' already exists")
                        elif if_exists == "replace":
                            duck.execute(f"DROP TABLE IF EXISTS {qtbl}")
                            tbl_exists = False
                    created = False
                    while True:
                        batch = stream_res.fetchmany(50000)
                        if not batch:
                            break
                        rows_as_lists = [[_to_json_safe(v) for v in row] for row in batch]
                        if not created:
                            sample = rows_as_lists[:64]
                            types: dict[str, str] = {c: "TEXT" for c in columns}
                            for idx, col in enumerate(columns):
                                for r in sample:
                                    try:
                                        v = r[idx]
                                        if v is None:
                                            continue
                                        if isinstance(v, bool):
                                            types[col] = "BOOLEAN"; break
                                        elif isinstance(v, int):
                                            types[col] = "BIGINT"; break
                                        elif isinstance(v, float):
                                            types[col] = "DOUBLE"; break
                                    except Exception:
                                        pass
                            cols_sql = ", ".join(f'{_q(c)} {types[c]}' for c in columns)
                            duck.execute(f"CREATE TABLE IF NOT EXISTS {qtbl} ({cols_sql})")
                            created = True
                        qmarks = ", ".join(["?" for _ in columns])
                        qcols = ", ".join([_q(c) for c in columns])
                        duck.executemany(
                            f"INSERT INTO {qtbl} ({qcols}) VALUES ({qmarks})",
                            rows_as_lists
                        )
                        total_rows += len(batch)
            return {"ok": True, "tableName": table_name, "rowCount": total_rows}
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Import failed: {e}")
    # --- Local DuckDB path: execute SQL directly on DuckDB ---
    try:
        with open_duck_native(duck_path) as conn:
            if if_exists == "replace":
                conn.execute(f"DROP TABLE IF EXISTS {qtbl}")
                conn.execute(f"CREATE TABLE {qtbl} AS ({sql})")
            elif if_exists == "append":
                exists = conn.execute(
                    "SELECT count(*) FROM information_schema.tables "
                    "WHERE table_schema='main' AND table_name=?",
                    [table_name],
                ).fetchone()[0]
                if exists:
                    conn.execute(f"INSERT INTO {qtbl} SELECT * FROM ({sql}) __q")
                else:
                    conn.execute(f"CREATE TABLE {qtbl} AS ({sql})")
            else:
                conn.execute(f"CREATE TABLE {qtbl} AS ({sql})")
            row_count = conn.execute(f"SELECT COUNT(*) FROM {qtbl}").fetchone()[0]
        return {"ok": True, "tableName": table_name, "rowCount": int(row_count)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Import failed: {e}")


@router.post("/{ds_id}/local/import-file-preview")
async def import_file_preview(
    ds_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Parse an uploaded CSV or Excel file and return preview rows."""
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    if _duckdb is None:
        raise HTTPException(status_code=500, detail="DuckDB driver unavailable")
    import tempfile, os as _os
    fname = (file.filename or "upload").lower()
    is_excel = fname.endswith(".xlsx") or fname.endswith(".xls")
    is_csv = fname.endswith(".csv") or fname.endswith(".tsv") or (not is_excel)
    content = await file.read()
    suffix = ".xlsx" if is_excel else ".csv"
    tmp = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
    try:
        tmp.write(content)
        tmp.flush()
        tmp.close()
        columns: list[str] = []
        rows: list[dict] = []
        if is_excel:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
                ws = wb.active
                raw_rows = list(ws.iter_rows(values_only=True))
                wb.close()
                if not raw_rows:
                    raise HTTPException(status_code=400, detail="Empty Excel file")
                columns = [str(c) if c is not None else f"col{i}" for i, c in enumerate(raw_rows[0])]
                import datetime, decimal
                def _safe(v):
                    if isinstance(v, (datetime.date, datetime.datetime)):
                        return str(v)
                    if isinstance(v, decimal.Decimal):
                        return float(v)
                    return v
                for r in raw_rows[1:101]:
                    rows.append({columns[i]: _safe(v) for i, v in enumerate(r)})
            except ImportError:
                raise HTTPException(status_code=500, detail="openpyxl not installed")
        else:
            # Use DuckDB's built-in CSV reader for fast parsing
            with open_duck_native(":memory:") as conn:
                res = conn.execute(f"SELECT * FROM read_csv_auto('{tmp.name}') LIMIT 100")
                columns = [str(c[0]) for c in (res.description or [])]
                raw = res.fetchall()
            import datetime, decimal
            def _safe(v):
                if isinstance(v, (datetime.date, datetime.datetime)):
                    return str(v)
                if isinstance(v, decimal.Decimal):
                    return float(v)
                return v
            rows = [{k: _safe(v) for k, v in zip(columns, row)} for row in raw]
        return {"columns": columns, "rows": rows, "rowCount": len(rows), "fileName": file.filename}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File parse error: {e}")
    finally:
        try:
            _os.unlink(tmp.name)
        except Exception:
            pass


@router.post("/{ds_id}/local/import-file-commit")
async def import_file_commit(
    ds_id: str,
    file: UploadFile = File(...),
    tableName: str = Form(...),
    ifExists: str = Form(default="replace"),
    db: Session = Depends(get_db),
):
    """Commit an uploaded CSV/Excel file as a new DuckDB table."""
    ds = db.get(Datasource, ds_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Datasource not found")
    if _duckdb is None:
        raise HTTPException(status_code=500, detail="DuckDB driver unavailable")
    table_name = (tableName or "").strip()
    if not table_name:
        raise HTTPException(status_code=400, detail="tableName is required")
    if_exists = str(ifExists or "replace").lower()
    fname = (file.filename or "upload").lower()
    is_excel = fname.endswith(".xlsx") or fname.endswith(".xls")
    content = await file.read()
    import tempfile, os as _os
    suffix = ".xlsx" if is_excel else ".csv"
    tmp = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
    duck_path = _duck_path_for_ds(ds)
    qtbl = _q(table_name)
    try:
        tmp.write(content)
        tmp.flush()
        tmp.close()
        with open_duck_native(duck_path) as conn:
            if is_excel:
                # Parse with openpyxl, write to a temp CSV, then load
                try:
                    import openpyxl, csv, io
                    wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
                    ws = wb.active
                    raw_rows = list(ws.iter_rows(values_only=True))
                    wb.close()
                    if not raw_rows:
                        raise HTTPException(status_code=400, detail="Empty Excel file")
                    tmp_csv = tempfile.NamedTemporaryFile(suffix=".csv", delete=False, mode="w", newline="", encoding="utf-8")
                    try:
                        writer = csv.writer(tmp_csv)
                        for row in raw_rows:
                            writer.writerow([("" if v is None else str(v)) for v in row])
                        tmp_csv.close()
                        if if_exists == "replace":
                            conn.execute(f"DROP TABLE IF EXISTS {qtbl}")
                            conn.execute(f"CREATE TABLE {qtbl} AS SELECT * FROM read_csv_auto('{tmp_csv.name}')")
                        elif if_exists == "append":
                            exists = conn.execute(
                                "SELECT count(*) FROM information_schema.tables WHERE table_schema='main' AND table_name=?",
                                [table_name],
                            ).fetchone()[0]
                            if exists:
                                conn.execute(f"INSERT INTO {qtbl} SELECT * FROM read_csv_auto('{tmp_csv.name}') OFFSET 1")
                            else:
                                conn.execute(f"CREATE TABLE {qtbl} AS SELECT * FROM read_csv_auto('{tmp_csv.name}')")
                        else:
                            conn.execute(f"CREATE TABLE {qtbl} AS SELECT * FROM read_csv_auto('{tmp_csv.name}')")
                    finally:
                        try:
                            _os.unlink(tmp_csv.name)
                        except Exception:
                            pass
                except ImportError:
                    raise HTTPException(status_code=500, detail="openpyxl not installed")
            else:
                if if_exists == "replace":
                    conn.execute(f"DROP TABLE IF EXISTS {qtbl}")
                    conn.execute(f"CREATE TABLE {qtbl} AS SELECT * FROM read_csv_auto('{tmp.name}')")
                elif if_exists == "append":
                    exists = conn.execute(
                        "SELECT count(*) FROM information_schema.tables WHERE table_schema='main' AND table_name=?",
                        [table_name],
                    ).fetchone()[0]
                    if exists:
                        conn.execute(f"INSERT INTO {qtbl} SELECT * FROM read_csv_auto('{tmp.name}')")
                    else:
                        conn.execute(f"CREATE TABLE {qtbl} AS SELECT * FROM read_csv_auto('{tmp.name}')")
                else:
                    conn.execute(f"CREATE TABLE {qtbl} AS SELECT * FROM read_csv_auto('{tmp.name}')")
            row_count = conn.execute(f"SELECT COUNT(*) FROM {qtbl}").fetchone()[0]
        return {"ok": True, "tableName": table_name, "rowCount": int(row_count)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Import failed: {e}")
    finally:
        try:
            _os.unlink(tmp.name)
        except Exception:
            pass
